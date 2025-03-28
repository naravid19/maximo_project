# path : maximo_project/maximo_app/views.py

# Standard Library Imports
import datetime
import io
import logging
import os
import re
import regex
import shutil
import sys
import time
import traceback
import uuid
import warnings

# Third-Party Imports
import numpy as np
import pandas as pd
import xlwings as xw
from background_task import background
from django.conf import settings
from django.contrib import messages
from django.http import FileResponse, HttpResponse, Http404, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.http import require_GET
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

# Local Imports
from .forms import UploadFileForm
from maximo_app.models import Site, ChildSite, PlantType, Unit, WorkType, ActType, WBSCode, Status

warnings.simplefilter("ignore")

# ตั้งค่า stdout ให้ใช้การเข้ารหัสแบบ utf-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
logging.basicConfig(
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(sys.stdout),  # แสดงผลในคอนโซลด้วย utf-8
        logging.FileHandler("logfile.log", encoding="utf-8"),  # บันทึกลงไฟล์ด้วย utf-8
    ]
)

logger = logging.getLogger(__name__)

# Create your views here.
# ---------------------------------
# ฟังก์ชันพื้นฐาน
# ---------------------------------
orgid = 'EGAT'
pluscrevum = 0
status = 'ACTIVE'
pluscjprevnum = 0
frequnit = 'YEARS'
leadtime = 7

# def test_404(request):
#     return render(request, "errors/404.html")

# def test_500(request):
#     return render(request, "errors/500.html")

def index(request):
    schedule_filename = None
    location_filename = None
    missing_messages = []
    invalid_messages = []
    error_messages = []
    selected_order = []

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            temp_dir = 'temp'
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            template_dir = os.path.join(settings.STATIC_ROOT, 'excel')
            if not os.path.exists(template_dir):
                os.makedirs(template_dir)
            
            schedule_file = request.FILES.get('schedule_file', None)
            location_file = request.FILES.get('location_file', None)
            year = form.cleaned_data.get('year')
            frequency = form.cleaned_data.get('frequency', '4')
            plant_type = form.cleaned_data.get('plant_type')
            site = form.cleaned_data.get('site')
            child_site = form.cleaned_data.get('child_site')
            unit = form.cleaned_data.get('unit')
            wostatus = form.cleaned_data.get('wostatus')
            work_type = form.cleaned_data.get('work_type')
            acttype = form.cleaned_data.get('acttype')
            wbs = form.cleaned_data.get('wbs')
            wbs_other = form.cleaned_data.get('wbs_other') if wbs.wbs_code == 'อื่นๆ' else None
            projectid_other = form.cleaned_data.get('projectid_other') if wbs.wbs_code == 'อื่นๆ' else None
            selected_order = form.cleaned_data.get('selected_order', [])
            log_params = []
            log_error = []
            
            if not schedule_file:
                log_error.append(f"Schedule file")
            
            if not location_file:
                log_error.append(f"Location file")
            
            if year:
                buddhist_year = int(year) + 543
                two_digits_year = buddhist_year % 100
                log_params.append(f"PLANT OUTAGE YEAR: {year}")
            else:
                log_error.append("PLANT OUTAGE YEAR")
            
            if frequency:
                log_params.append(f"FREQUENCY: {frequency}")
            else:
                log_error.append("FREQUENCY")
            
            if plant_type:
                plant_type = plant_type.plant_code
                log_params.append(f"PLANT TYPE: {plant_type}")
            else:
                log_error.append("PLANT TYPE")
            
            if site:
                siteid = site.site_id
                log_params.append(f"SITE: {siteid}")
            else:
                log_error.append("SITE")
            
            if child_site:
                child_site = child_site.site_id
                log_params.append(f"PLANT NAME: {child_site}")
            else:
                log_error.append("PLANT NAME")
            
            if unit:
                unit = unit.unit_code
                if plant_type == 'PV' and unit.isdigit():
                    unit = str(int(unit))
                log_params.append(f"UNIT: {unit}")
            else:
                log_error.append("UNIT")
            
            if wostatus:
                wostatus = wostatus.status
                log_params.append(f"STATUS: {wostatus}")
            else:
                log_error.append("STATUS")
            
            if work_type:
                worktype = work_type.worktype
                log_params.append(f"WORK TYPE: {worktype}")
            else:
                log_error.append("WORK TYPE")
            
            if acttype:
                egmntacttype = acttype.acttype
                log_params.append(f"MNTACT TYPE: {egmntacttype}")
            else:
                log_error.append("MNTACT TYPE")
            
            if wbs and wbs.wbs_code != 'อื่นๆ':
                log_params.append(f"SUBWBS GROUP: {wbs.wbs_code}")
            elif wbs.wbs_code == 'อื่นๆ':
                if wbs_other:
                        log_params.append(f"Other WBS: {wbs_other}")
                else:
                    log_error.append("Other WBS")
            else:
                log_error.append("SUBWBS GROUP")
            
            if selected_order:
                if isinstance(selected_order, str):
                    selected_order = [selected_order]
                log_params.append(f"GROUPING: {selected_order}")
                grouping_text = get_grouping_text(selected_order)
            else:
                log_error.append("GROUPING")
            
            # if log_params:
            #     logger.info("\n".join(log_params))
            
            if log_error:
                for error in log_error:
                    logger.error(f"'{error}' was not provided from the form.")
                error_message = (
                    f"<div class='error-container'>"
                    f"<strong class='error-title'>ข้อผิดพลาด:</strong> ข้อมูลที่จำเป็นไม่ได้รับการระบุจากฟอร์ม<br>"
                    f"<ul class='error-details'>"
                    f"{''.join(f'<li>{error}</li>' for error in log_error)}"
                    f"</ul>"
                    f"<p class='error-note'>*** โปรดตรวจสอบและเลือกข้อมูลที่จำเป็นในฟอร์มทุกฟิลด์ที่เกี่ยวข้อง ***</p>"
                    f"</div>"
                )
                messages.error(request, error_message)
                return redirect('index')
            
            try:
                location = f'{child_site}-{plant_type}{unit}'
                location_sanitized = location.replace('-', '')
                
                if wbs and wbs.wbs_code != 'อื่นๆ':
                    egprojectid = f"O-{location_sanitized}-{two_digits_year}{acttype.code}"
                    egwbs = f"{egprojectid}-{wbs.wbs_code}"
                elif wbs.wbs_code == 'อื่นๆ' and wbs_other:
                    egwbs = wbs_other
                    egprojectid = projectid_other
                else:
                    egwbs = ''
                    egprojectid = ''
                
                wbs_desc = f"{wbs.description} {acttype.description} {location} {buddhist_year}"
                
            except Exception as e:
                logger.error(f"An error occurred: {str(e)}", exc_info=True)
                error_message = (
                    f"<div class='error-container'>"
                    f"<strong class='error-title'>พบปัญหา:</strong> เกิดข้อขัดข้องระหว่างการประมวลผลข้อมูล<br>"
                    f"<ul class='error-details'>"
                    f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                    f"<li>{str(e)}</li>"
                    f"</ul>"
                    f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบข้อมูลที่ป้อนและลองดำเนินการอีกครั้ง หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                    f"</div>"
                )
                messages.error(request, error_message)
                return redirect('index')
            
            schedule_filename = schedule_file.name
            location_filename = location_file.name
            unique_schedule_name = f"{uuid.uuid4()}_{schedule_file.name}"
            unique_location_name = f"{uuid.uuid4()}_{location_file.name}"
            schedule_path = os.path.join(temp_dir, unique_schedule_name)
            location_path = os.path.join(temp_dir, unique_location_name)
            comment_path = os.path.join(temp_dir, f"{uuid.uuid4()}_Comment.xlsx")
            
            try:
                with open(schedule_path, 'wb+') as destination:
                    for chunk in schedule_file.chunks():
                        destination.write(chunk)

                with open(location_path, 'wb+') as destination:
                    for chunk in location_file.chunks():
                        destination.write(chunk)
            
            except IOError as e:
                logger.error(f"Error saving files: {str(e)}")
                error_message = (
                    f"<div class='error-container'>"
                    f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถบันทึกไฟล์ได้<br>"
                    f"<ul class='error-details'>"
                    f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                    f"<li>{str(e)}</li>"
                    f"</ul>"
                    f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบสิ่งต่อไปนี้<br>"
                    f"&nbsp;&nbsp;&nbsp;&nbsp;1. ตำแหน่งที่บันทึกไฟล์สามารถเข้าถึงได้<br>"
                    f"&nbsp;&nbsp;&nbsp;&nbsp;2. คุณมีสิทธิ์ในการบันทึกไฟล์ที่ตำแหน่งนี้<br>"
                    f"&nbsp;&nbsp;&nbsp;&nbsp;3. มีพื้นที่เพียงพอสำหรับบันทึกไฟล์<br>"
                    f"หากยังไม่สามารถแก้ไขได้ กรุณาติดต่อทีมสนับสนุนพร้อมแจ้งข้อความข้างต้น"
                    f"</p>"
                    f"</div>"
                )
                messages.error(request, error_message)
                return redirect('index')

            request.session['schedule_filename'] = schedule_file.name
            request.session['location_filename'] = location_file.name
            request.session['temp_dir'] = temp_dir
            request.session['schedule_path'] = schedule_path
            request.session['location_path'] = location_path
            request.session['year'] = year
            request.session['frequency'] = frequency
            request.session['egmntacttype'] = egmntacttype
            request.session['egprojectid'] = egprojectid
            request.session['egwbs'] = egwbs
            request.session['location'] = location
            request.session['siteid'] = siteid
            request.session['child_site'] = child_site
            request.session['wostatus'] = wostatus
            request.session['wbs'] = wbs.wbs_code
            request.session['wbs_desc'] = wbs_desc
            request.session['worktype'] = worktype
            request.session['wostatus'] = wostatus
            request.session['grouping_text'] = grouping_text

            try:
                required_columns = [
                    'KKS', 'EQUIPMENT', 'TASK_XX', 'TASK', 
                    'RESPONSE', 'ROUTE', 'DURATION_(HR.)', 'START_DATE', 'FINISH_DATE', 
                    'SUPERVISOR', 'FOREMAN', 'SKILL', 'RESPONSE_CRAFT', 
                    'ประเภทของ_PERMIT_TO_WORK', 'TYPE', 'COMMENT'
                ]
                
                important_columns = ['TASK_XX']
                
                use_columns = [
                    'KKS', 'EQUIPMENT', 'TASK_XX', 'TASK', 
                    'RESPONSE', 'ROUTE', 'DURATION_(HR.)', 'START_DATE', 'FINISH_DATE', 
                    'SUPERVISOR', 'FOREMAN', 'SKILL', 'RESPONSE_CRAFT', 
                    'ประเภทของ_PERMIT_TO_WORK', 'TYPE'
                ]
                
                location_columns = ['Location', 'Description']
                
                df_original = read_excel_with_error_handling(request, schedule_path)
                if df_original is None:
                    raise ValueError("Cannot proceed without a valid DataFrame.")
                # logger.info("Excel file loaded successfully.")
                
                df_original.columns = [col.strip().upper().replace(' ', '_') if isinstance(col, str) else col for col in df_original.columns]
                # logger.info("Column names cleaned and formatted.")
                
                # ตรวจสอบว่าคอลัมน์ที่จำเป็นทั้งหมดมีอยู่หรือไม่
                missing_columns = [col for col in required_columns if col not in df_original.columns]
                if missing_columns:
                    logger.error(f"Missing required columns: {', '.join(missing_columns)}")
                    error_message = (
                        f"<div class='error-container'>"
                        f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบคอลัมน์ที่จำเป็น<br>"
                        f"<ul class='error-details'>"
                        f"<p class='error-description'>ไฟล์ {schedule_filename} ขาดคอลัมน์ต่อไปนี้:</p>"
                        f"{''.join(f'<li>{col}</li>' for col in missing_columns)}"
                        f"</ul>"
                        f"<p class='error-note'>คำแนะนำ:</p>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;1. โปรดตรวจสอบว่ามีคอลัมน์ที่ระบุข้างต้นหรือไม่<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;2. หากไม่มี ให้เพิ่มคอลัมน์ที่ขาดและกรอกข้อมูลให้ครบถ้วน<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;3. บันทึกไฟล์และอัปโหลดใหม่อีกครั้ง"
                        f"</p>"
                        f"</div>"
                    )
                    messages.error(request, error_message)
                    return redirect('index')
                
                # ตรวจสอบว่าคอลัมน์ที่สำคัญมีข้อมูลอย่างน้อยหนึ่งค่า
                empty_columns = [col for col in important_columns if df_original[col].isna().all() or (df_original[col] == '').all()]
                if empty_columns:
                    logger.error(f"Important columns without data: {', '.join(empty_columns)}")
                    error_message = (
                        f"<div class='error-container'>"
                        f"<strong class='error-title'>พบปัญหา: คอลัมน์ไม่มีข้อมูล</strong> <br>"
                        f"<ul class='error-details'>"
                        f"<p class='error-description'>คอลัมน์ต่อไปนี้ไม่มีข้อมูล:</p>"
                        f"{''.join(f'<li>{col}</li>' for col in empty_columns)}"
                        f"</ul>"
                        f"<p class='error-note'>คำแนะนำ:</p>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;1. ตรวจสอบคอลัมน์ที่ระบุข้างต้น<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;2. หากไม่มีข้อมูลเลย ให้เพิ่มข้อมูลที่จำเป็นลงในคอลัมน์เหล่านั้น<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;3. ตรวจสอบให้แน่ใจว่าไม่มีคอลัมน์ใดว่างเปล่าทั้งหมด<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;4. บันทึกไฟล์และอัปโหลดใหม่อีกครั้ง<br>"
                        f"</p>"
                        f"</div>"
                    )
                    messages.error(request, error_message)
                    return redirect('index')
                
                kks_read = pd.read_excel(location_file, 
                                        sheet_name=0,header = 0,
                                        usecols = 'A:B',
                                        engine="openpyxl")
                
                location_column_errors = [col for col in location_columns if col not in kks_read.columns]
                if location_column_errors:
                    logger.error(f"Missing columns in Location file: {', '.join(location_column_errors)}")
                    error_message = (
                        f"<div class='error-container'>"
                        f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบคอลัมน์ที่จำเป็นในไฟล์ Location<br>"
                        f"<ul class='error-details'>"
                        f"<p class='error-description'>ไฟล์ {location_filename} ขาดคอลัมน์ต่อไปนี้:</p>"
                        f"{''.join(f'<li>{col}</li>' for col in location_column_errors)}"
                        f"</ul>"
                        f"<p class='error-note'>คำแนะนำ:</p>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;1. โปรดตรวจสอบว่ามีคอลัมน์ที่ระบุข้างต้นหรือไม่<br>"
                        f"&nbsp;&nbsp;&nbsp;&nbsp;2. โปรดตรวจสอบว่าไฟล์ที่อัปโหลดเป็นไฟล์ Location<br>"
                        f"</p>"
                        f"</div>"
                    )
                    messages.error(request, error_message)
                    return redirect('index')
                
                df_original = df_original[use_columns]
                df_original.rename(columns={'TASK_XX': 'TASK_ORDER'}, inplace=True)
                use_columns = [col.replace('TASK_XX', 'TASK_ORDER') for col in use_columns]
                df_original['KKS'] = df_original['KKS'].apply(lambda x: x.upper() if isinstance(x, str) else x)
                df_original['ROUTE'] = df_original['ROUTE'].apply(lambda x: x.upper() if isinstance(x, str) else x)
                df_original['RESPONSE'] = df_original['RESPONSE'].apply(lambda x: x.upper() if isinstance(x, str) else x)
                df_original['RESPONSE_CRAFT'] = df_original['RESPONSE_CRAFT'].apply(lambda x: x.upper() if isinstance(x, str) else x)
                df_original['TYPE'] = df_original['TYPE'].apply(lambda x: x.upper() if isinstance(x, str) else x)
                
                for col in use_columns:
                    if col in df_original.columns:
                        df_original[col] = df_original[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
                
            except Exception as e:
                logger.error(f"An error occurred: {str(e)}", exc_info=True)
                error_message = (
                    f"<div class='error-container'>"
                    f"<strong class='error-title'>พบปัญหา:</strong> เกิดข้อขัดข้องระหว่างการดำเนินการ<br>"
                    f"<ul class='error-details'>"
                    f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                    f"<li>{str(e)}</li>"
                    f"<li>กรุณาตรวจสอบว่ารูปแบบของ Template Excel ถูกต้อง</li>"
                    f"</ul>"
                    f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                    f"</div>"
                )
                messages.error(request, error_message)
                return redirect('index')
            
            # logger.info("DataFrame preparation completed successfully.")
            
            df_comment = pd.DataFrame()
            df_comment['KKS'] = df_original['KKS']
            df_comment['COMMENT'] = ''

            df_original = df_original.replace('', np.nan)

            long_equipment = df_original['EQUIPMENT'].astype(str).str.len() > 100
            update_comment(df_comment, long_equipment, 'COMMENT', 'EQUIPMENT มีความยาวมากกว่า 100 ตัวอักษร')

            task_with_order = ((df_original['TASK_ORDER'].notna())) & ((df_original['TASK_ORDER']!='xx'))
            long_task = df_original['TASK'].astype(str).str.len() > 100
            update_comment(df_comment, long_task, 'COMMENT', 'TASK มีความยาวมากกว่า 100 ตัวอักษร')

            task_na = df_original['TASK'].isna()
            update_comment(df_comment, (task_with_order & task_na), 'COMMENT', 'ไม่มี TASK')
            
            route_length_exceed = df_original['ROUTE'].astype(str).str.len() > 20
            update_comment(df_comment, route_length_exceed, 'COMMENT', 'ROUTE มีความยาวมากกว่า 20 ตัวอักษร')
            
            df_original['DURATION_(HR.)'] = df_original['DURATION_(HR.)'].apply(convert_duration)

            # ตรวจสอบค่า DURATION_(HR.) ว่าเป็นตัวเลขหรือไม่
            cond_duration = df_original['DURATION_(HR.)'].apply(lambda x: isinstance(x, (int, float)))
            non_negative = df_original[cond_duration]['DURATION_(HR.)'] < 0
            task_order_not_xx = df_original['TASK_ORDER'] != 'xx'
            update_comment(df_comment, ((~cond_duration | non_negative) & task_order_not_xx), 'COMMENT', 'DURATION_(HR.) ไม่ถูกต้อง')

            cond_start_date = df_original['START_DATE'].apply(lambda x: isinstance(x, str) and regex.search(r"\p{L}", x) is not None and not is_date(x))
            update_comment(df_comment, cond_start_date, 'COMMENT', 'START_DATE มีตัวอักษร')

            cond_finish_date = df_original['FINISH_DATE'].apply(lambda x: isinstance(x, str) and regex.search(r"\p{L}", x) is not None and not is_date(x))
            update_comment(df_comment, cond_finish_date, 'COMMENT', 'FINISH_DATE มีตัวอักษร')

            df_original['START_DATE_NEW'] = parse_dates(df_original['START_DATE'])
            df_original['FINISH_DATE_NEW'] = parse_dates(df_original['FINISH_DATE'])

            cond_start_date_new = df_original['START_DATE'].notna() & df_original['START_DATE_NEW'].isna() & (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx')
            update_comment(df_comment, cond_start_date_new, 'COMMENT', 'START_DATE ไม่ถูกต้อง')

            cond_finish_date_new = df_original['FINISH_DATE'].notna() & df_original['FINISH_DATE_NEW'].isna() & (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx')
            update_comment(df_comment, cond_finish_date_new, 'COMMENT', 'FINISH_DATE ไม่ถูกต้อง')

            kks_length_exceed = df_original['KKS'].astype(str).str.len() > 30
            update_comment(df_comment, kks_length_exceed, 'COMMENT', 'KKS มีความยาวมากกว่า 30 ตัวอักษร')

            # หาค่า Plant Unit ที่พบมากที่สุด
            most_common_plant_unit = df_original['KKS'].str.split('-', expand=True)[0].value_counts().idxmax()
            # สร้างเงื่อนไขสำหรับค่า Plant Unit ที่ไม่ตรงกับค่าที่พบมากที่สุด
            cond_kks = (df_original['KKS'].str.split('-', expand=True)[0] != most_common_plant_unit) & (df_original['KKS'].notna())
            update_comment(df_comment, cond_kks, 'COMMENT', 'Plant Unit ไม่สอดคล้อง')

            plant_list = df_original['KKS'].str.split('-',expand =True)[0].value_counts().index.tolist()
            plant_regex = "|".join([p + "-" for p in plant_list])
            first_plant = plant_regex.split('|')[0]
            df_original['KKS'] = df_original['KKS'].replace(plant_regex,'', regex=True)

            # cond1 = (df_original['TASK_ORDER']==10) & ((df_original['DURATION (HR.)'].isna())|(df_original['START DATE_new'].isna())|(df_original['FINISH DATE_new'].isna()))
            # cond2 = (df_original['TASK_ORDER']==10) & ((df_original['SUPERVISOR'].isna())&(df_original['FOREMAN'].isna())&(df_original['SKILL'].isna())&(df_original['RESPONSE CRAFT'].isna()))
            # cond3 = (df_original['TASK_ORDER']==10) & (df_original['SUPERVISOR'].isna())&(df_original['FOREMAN'].isna())&(df_original['SKILL'].isna())
            task_no_start_date = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['START_DATE_NEW'].isna()) & (df_original['START_DATE'].isna())
            task_no_finish_date = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['FINISH_DATE_NEW'].isna()) & (df_original['FINISH_DATE'].isna())
            task_no_skill_rate = (df_original['TASK_ORDER'] == 10) & ((df_original['SUPERVISOR'].isna()) & (df_original['FOREMAN'].isna()) & (df_original['SKILL'].isna()))
            task_no_craft = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['RESPONSE_CRAFT'].isna())
            task_no_response = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['RESPONSE'].isna())

            update_comment(df_comment, task_no_start_date, 'COMMENT', 'ไม่มี START_DATE')
            update_comment(df_comment, task_no_finish_date, 'COMMENT', 'ไม่มี FINISH_DATE')
            update_comment(df_comment, task_no_skill_rate, 'COMMENT', 'ไม่มี SKILL RATE (จำเป็นต้องกรอก)')

            task_order = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx')
            # ตรวจสอบว่า SUPERVISOR, FOREMAN, SKILL ถ้ามีค่า (ไม่ใช่ NaN) ต้องเป็นจำนวนเต็มบวก
            valid_supervisor = df_original['SUPERVISOR'].apply(lambda x: (pd.isna(x) or (isinstance(x, (int, float)) and x.is_integer() and x >= 0)))
            valid_foreman = df_original['FOREMAN'].apply(lambda x: (pd.isna(x) or (isinstance(x, (int, float)) and x.is_integer() and x >= 0)))
            valid_skill = df_original['SKILL'].apply(lambda x: (pd.isna(x) or (isinstance(x, (int, float)) and x.is_integer() and x >= 0)))
            # เงื่อนไขตรวจสอบว่าค่าใน SUPERVISOR, FOREMAN, SKILL มีค่าที่ไม่ถูกต้อง (ไม่เป็นจำนวนเต็มบวก)
            invalid_skill_rate = ~(valid_supervisor & valid_foreman & valid_skill)
            update_comment(df_comment, (task_order & invalid_skill_rate), 'COMMENT', 'SKILL RATE ไม่ถูกต้อง')
            update_comment(df_comment, task_no_craft, 'COMMENT', 'ไม่มี RESPONSE_CRAFT')
            carft_length_exceed = df_original['RESPONSE_CRAFT'].astype(str).str.len() > 12
            update_comment(df_comment, carft_length_exceed, 'COMMENT', 'RESPONSE_CRAFT มีความยาวมากกว่า 12 ตัวอักษร')
            update_comment(df_comment, task_no_response, 'COMMENT', 'ไม่มี RESPONSE')
            response_length_exceed = df_original['RESPONSE'].astype(str).str.len() > 12 
            update_comment(df_comment, response_length_exceed, 'COMMENT', 'RESPONSE มีความยาวมากกว่า 12 ตัวอักษร')

            df_original_copy = df_original.copy()
            df_original_copy = df_original_copy.dropna(how='all')
            df_original_copy = df_original_copy.reset_index(drop=True)
            df_original_copy = df_original_copy.fillna(-1)

            temp_kks = ''
            lst_kks = []

            temp_eq = ''
            lst_eq = []

            temp_task_order = ''
            lst_task_order = []

            temp_task = ''
            lst_task = []

            temp_start_date = ''
            lst_start_date = []

            temp_finish_date = ''
            lst_finish_date = []

            temp_type_work = ''
            lst_type_work = []

            temp_response = ''
            lst_response = []

            temp_RESPONSE_CRAFT = ''
            lst_RESPONSE_CRAFT = []

            for i in (df_original_copy['KKS']):
                if i!=-1:
                    temp_kks = i
                    lst_kks.append(i)
                else:
                    i=temp_kks
                    lst_kks.append(i)

            for j in (df_original_copy['EQUIPMENT']):
                if j!=-1:
                    temp_eq = j
                    lst_eq.append(j)
                else:
                    j=temp_eq
                    lst_eq.append(j)

            for k in (df_original_copy['TASK_ORDER']):
                if k!=-1:
                    temp_task_order = k
                    lst_task_order.append(k)
                else:
                    k=temp_task_order
                    lst_task_order.append(k)

            for z in (df_original_copy['TASK']):
                if z!=-1:
                    temp_task = z
                    lst_task.append(z)
                else:
                    z=temp_task
                    lst_task.append(z)

            for xx in (df_original_copy['START_DATE_NEW']):
                if xx!=-1:
                    temp_start_date = xx
                    lst_start_date.append(xx)
                else:
                    xx=temp_start_date
                    lst_start_date.append(xx)

            for yy in (df_original_copy['FINISH_DATE_NEW']):
                if yy!=-1:
                    temp_finish_date = yy
                    lst_finish_date.append(yy)
                else:
                    yy=temp_finish_date
                    lst_finish_date.append(yy)

            for zz in (df_original_copy['ประเภทของ_PERMIT_TO_WORK']):
                if zz!=-1:
                    temp_type_work = zz
                    lst_type_work.append(zz)
                else:
                    zz=temp_type_work
                    lst_type_work.append(zz)

            for xy in (df_original_copy['RESPONSE']):
                if xy!=-1:
                    temp_response = xy
                    lst_response.append(xy)
                else:
                    xy=temp_response
                    lst_response.append(xy)

            for xy in (df_original_copy['RESPONSE_CRAFT']):
                if xy!=-1:
                    temp_RESPONSE_CRAFT = xy
                    lst_RESPONSE_CRAFT.append(xy)
                else:
                    xy=temp_RESPONSE_CRAFT
                    lst_RESPONSE_CRAFT.append(xy)

            df_original_copy['KKS_NEW'] = pd.Series(lst_kks)
            df_original_copy['EQUIPMENT_NEW'] = pd.Series(lst_eq)
            df_original_copy['TASK_ORDER_NEW'] = pd.Series(lst_task_order)
            df_original_copy['TASK_NEW']= pd.Series(lst_task)

            df_original_copy['START_DATE']=pd.Series(lst_start_date)
            df_original_copy['FINISH_DATE']=pd.Series(lst_finish_date)
            df_original_copy['ประเภทของ_PERMIT_TO_WORK'] = pd.Series(lst_type_work)
            df_original_copy['RESPONSE'] = pd.Series(lst_response)
            df_original_copy['RESPONSE_CRAFT'] = pd.Series(lst_RESPONSE_CRAFT)
            
            for idx in range(len(df_original_copy) - 1):
                if df_original_copy.loc[idx, 'TASK_ORDER'] == 'xx' and df_original_copy.loc[idx, 'ROUTE'] != -1:
                    next_idx = idx + 1

                    if df_original_copy.loc[next_idx, 'TASK_ORDER'] != 'xx' and df_original_copy.loc[next_idx, 'ROUTE'] == -1:
                        df_original_copy.loc[next_idx, 'ROUTE'] = df_original_copy.loc[idx, 'ROUTE']

            lst = ['KKS_NEW', 'EQUIPMENT_NEW', 'ROUTE', 'TASK_ORDER_NEW', 'TASK_NEW',
                    'RESPONSE', 'DURATION_(HR.)', 'START_DATE', 'FINISH_DATE', 'SUPERVISOR',
                    'FOREMAN', 'SKILL', 'RESPONSE_CRAFT', 'ประเภทของ_PERMIT_TO_WORK','TYPE']
            df_original_newcol = df_original_copy.loc[:,lst].copy()
            main_system = df_original_newcol['KKS_NEW'].str[0:6]
            sub_system =  df_original_newcol['KKS_NEW'].str[0:8]
            equipment =   df_original_newcol['KKS_NEW'].str[8:10]

            df_original_newcol['MAIN_SYSTEM'] = main_system
            df_original_newcol['SUB_SYSTEM'] = sub_system
            df_original_newcol['EQUIPMENT'] = equipment

            plant_list1 = kks_read['Location'].str.split('-',expand=True)[0].value_counts().index.tolist()
            plant_regex1 = "|".join([p + "-" for p in plant_list1])
            kks_read['Location_x'] = kks_read['Location'].str.replace(plant_regex1, '', regex=True)

            lst_main_sys = []
            for i,j in df_original_newcol.iterrows():
                if len(kks_read[kks_read['Location_x']==j['MAIN_SYSTEM']]['Description'])==1:
                    desc = kks_read[kks_read['Location_x']==j['MAIN_SYSTEM']]['Description'].values[0]
                    lst_main_sys.append(desc)
                else:
                    lst_main_sys.append('No_kks_found')

            lst_sub_sys = []
            for i,j in df_original_newcol.iterrows():
                if len(kks_read[kks_read['Location_x']==j['SUB_SYSTEM']]['Description'])==1:
                    desc = kks_read[kks_read['Location_x']==j['SUB_SYSTEM']]['Description'].values[0]
                    lst_sub_sys.append(desc)
                else:
                    lst_sub_sys.append('No_kks_found')


            lst_equipment = []
            for i,j in df_original_newcol.iterrows():
                if len(kks_read[kks_read['Location_x']==j['KKS_NEW']]['Description'])==1:
                    desc = kks_read[kks_read['Location_x']==j['KKS_NEW']]['Description'].values[0]
                    lst_equipment.append(desc)
                else:
                    lst_equipment.append('No_kks_found')

            main_sys = pd.Series(lst_main_sys)
            sub_sys = pd.Series(lst_sub_sys)
            eq_sys = pd.Series(lst_equipment)
            df_original_newcol['MAIN_SYSTEM_DESC']=main_sys
            df_original_newcol['SUB_SYSTEM_DESC']=sub_sys
            df_original_newcol['KKS_NEW_DESC']=eq_sys

            cond1 = df_original_newcol['KKS_NEW'].notna()
            cond2 = df_original_newcol['MAIN_SYSTEM_DESC']=='No_kks_found'
            cond3 = df_original_newcol['SUB_SYSTEM_DESC']=='No_kks_found'
            cond4 = df_original_newcol['KKS_NEW_DESC']=='No_kks_found'
            cond5 = df_original_newcol['KKS_NEW'].isin([''])
            cond6 = df_original_newcol['TASK_ORDER_NEW']!='xx'

            update_comment(df_comment, (cond1 & cond4 & ~cond5 & ((df_original_copy['KKS']!= -1))), 'COMMENT', 'ไม่พบ kks')

            task_no_duration = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['DURATION_(HR.)'].isna())
            update_comment(df_comment, task_no_duration, 'COMMENT', 'ไม่มี DURATION_(HR.)')
            
            duration_length_exceed = df_original['DURATION_(HR.)'].apply(
                lambda x: len(f'{int(x):.0f}') > 8 if pd.notna(x) and isinstance(x, (int, float)) else False
            )
            update_comment(df_comment, duration_length_exceed, 'COMMENT', 'DURATION_(HR.) มีความยาวมากกว่า 8 หลัก')

            task_no_ptw = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['ประเภทของ_PERMIT_TO_WORK'].isna())
            update_comment(df_comment, task_no_ptw, 'COMMENT', 'ไม่มี ประเภทของ_PERMIT_TO_WORK')

            ptw_length_exceed = df_original['ประเภทของ_PERMIT_TO_WORK'].astype(str).str.len() > 250
            update_comment(df_comment, ptw_length_exceed, 'COMMENT', 'ประเภทของ_PERMIT_TO_WORK มีความยาวมากกว่า 250 ตัวอักษร')

            task_no_type = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['TYPE'].isna())
            update_comment(df_comment, task_no_type, 'COMMENT', 'ไม่มี TYPE (จำเป็นต้องกรอก)')

            task_type = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx') & (df_original['TYPE'].notna())
            valid_type = (df_original['TYPE'].isin(['ME', 'EE', 'CV', 'IC']))
            update_comment(df_comment, (task_type & ~valid_type), 'COMMENT', 'TYPE ไม่ถูกต้อง')

            df_original_check = df_original_copy.copy()
            df_original_check['TASK_ORDER_NEW'] = df_original_check['TASK_ORDER_NEW'].astype(str)
            # ตรวจสอบว่าค่าใน TASK_ORDER_NEW เป็นจำนวนเต็มบวกหรือ 'xx' เท่านั้น
            task_order_valid = df_original_check['TASK_ORDER_NEW'].apply(
                lambda x: (isinstance(x, str) and (x.isdigit() or x == 'xx')) or 
                        (isinstance(x, (int, float)) and x.is_integer() and x >= 0)
            )

            update_comment(df_comment, (~task_order_valid & (df_original_check['TASK_ORDER']!= -1)), 'COMMENT', 'TASK_ORDER ไม่ถูกต้อง')

            duration_notna_cond = df_original['DURATION_(HR.)'].notna()
            task_missing_cond = df_original['TASK_ORDER'].isna() | (df_original['TASK_ORDER'] == '')
            update_comment(df_comment, (duration_notna_cond & task_missing_cond), 'COMMENT', 'ไม่มี TASK_ORDER')

            task_order_length_exceed = df_original_check['TASK_ORDER_NEW'].astype(str).str.len() > 12
            update_comment(df_comment, (task_order_length_exceed & (df_original_check['TASK_ORDER']!= -1)), 'COMMENT', 'TASK_ORDER มีความยาวมากกว่า 12 ตัวอักษร')

            task_not_xx = df_original_copy['TASK_ORDER'] != 'xx'
            kks_new_na = df_original_copy['KKS_NEW'].isin([''])
            equip_new_na = df_original_copy['EQUIPMENT_NEW'].isin([''])
            task_order_new_na = df_original_copy['TASK_ORDER_NEW'].isin([''])
            task_new_na = df_original_copy['TASK_NEW'].isin([''])
            start_date_na = df_original_copy['START_DATE'].isin([''])
            finish_date_na = df_original_copy['FINISH_DATE'].isin([''])
            ptw_na = df_original_copy['ประเภทของ_PERMIT_TO_WORK'].isin([''])
            response_new_na = df_original_copy['RESPONSE'].isin([''])
            response_craft_na = df_original_copy['RESPONSE_CRAFT'].isin([''])
            
            replace_or_append_comment(df_comment, (task_not_xx & kks_new_na), 'COMMENT', 'ไม่มี KKS (จำเป็นต้องกรอก)', replace_message='ไม่มี KKS')
            update_comment(df_comment, (task_not_xx & equip_new_na), 'COMMENT', 'ไม่มี EQUIPMENT (จำเป็นต้องกรอก)')
            replace_or_append_comment(df_comment, (task_not_xx & task_order_new_na), 'COMMENT', 'ไม่มี TASK_ORDER (จำเป็นต้องกรอก)', replace_message='ไม่มี TASK_ORDER')
            replace_or_append_comment(df_comment, (task_not_xx & task_new_na), 'COMMENT', 'ไม่มี TASK (จำเป็นต้องกรอก)', replace_message='ไม่มี TASK')
            replace_or_append_comment(df_comment, (task_not_xx & start_date_na), 'COMMENT', 'ไม่มี START_DATE (จำเป็นต้องกรอก)', replace_message='ไม่มี START_DATE')
            replace_or_append_comment(df_comment, (task_not_xx & finish_date_na), 'COMMENT', 'ไม่มี FINISH_DATE (จำเป็นต้องกรอก)', replace_message='ไม่มี FINISH_DATE')
            replace_or_append_comment(df_comment, (task_not_xx & ptw_na), 'COMMENT', 'ไม่มี ประเภทของ_PERMIT_TO_WORK (จำเป็นต้องกรอก)', replace_message='ไม่มี ประเภทของ_PERMIT_TO_WORK')
            replace_or_append_comment(df_comment, (task_not_xx & response_new_na), 'COMMENT', 'ไม่มี RESPONSE (จำเป็นต้องกรอก)', replace_message='ไม่มี RESPONSE')
            replace_or_append_comment(df_comment, (task_not_xx & response_craft_na), 'COMMENT', 'ไม่มี RESPONSE_CRAFT (จำเป็นต้องกรอก)', replace_message='ไม่มี RESPONSE_CRAFT')

            df_comment['COMMENT'] = df_comment['COMMENT'].str.strip(', ')


            #! Create Commnet.xlsx file
            df_comment['TASK_ORDER'] = df_original_copy['TASK_ORDER_NEW']
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
            no_fill = PatternFill(fill_type=None)
            center_alignment = Alignment(horizontal='center')

            df_comment = df_comment[['TASK_ORDER','COMMENT']]
            book = load_workbook(schedule_path)
            sheet = book.worksheets[0]
            # sheet = book[sheet_name]
            start_row = 3
            task_order_col = 7  # คอลัมน์เริ่มต้นสำหรับ TASK_ORDER
            comment_col = 19

            for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
                if col[1].value and isinstance(col[1].value, str) and col[1].value.strip().upper() == "COMMENT":
                    comment_col = idx
                    break

            for i, (task_order, comment) in enumerate(zip(df_comment['TASK_ORDER'], df_comment['COMMENT']), start=start_row):
                # task_order_cell = sheet.cell(row=i, column=task_order_col)
                # task_order_cell.value = task_order
                # task_order_cell.alignment = center_alignment
                # task_order_cell.fill = blue_fill
                
                comment_cell = sheet.cell(row=i, column=comment_col)
                comment_cell.value = comment
                
                if comment_cell.value:
                    if "(จำเป็นต้องกรอก)" in comment_cell.value or "ไม่ถูกต้อง" in comment_cell.value:
                        comment_cell.fill = red_fill
                    else:
                        comment_cell.fill = yellow_fill
                else:
                    comment_cell.fill = no_fill
            
            book.save(comment_path)
            request.session['download_link_comment'] = comment_path
            
            #! RECHECK
            task_order_not_xx = df_original['TASK_ORDER'] != 'xx'
            task_order = (df_original['TASK_ORDER'].notna()) & (df_original['TASK_ORDER'] != 'xx')
            task_not_xx = df_original_copy['TASK_ORDER'] != 'xx'
            kks_new_na = df_original_copy['KKS_NEW'].isin([''])
            equip_new_na = df_original_copy['EQUIPMENT_NEW'].isin([''])
            task_order_new_na = df_original_copy['TASK_ORDER_NEW'].isin([''])
            task_new_na = df_original_copy['TASK_NEW'].isin([''])
            start_date_na = df_original_copy['START_DATE'].isin([''])
            finish_date_na = df_original_copy['FINISH_DATE'].isin([''])
            ptw_na = df_original_copy['ประเภทของ_PERMIT_TO_WORK'].isin([''])
            response_new_na = df_original_copy['RESPONSE'].isin([''])
            response_craft_na = df_original_copy['RESPONSE_CRAFT'].isin([''])

            # ตรวจสอบเงื่อนไขแต่ละคอลัมน์และเก็บชื่อคอลัมน์ที่ข้อมูลขาดหาย
            all_missing_columns = []
            missing_counts = []
            missing_indices = []

            if ((task_not_xx) & (kks_new_na)).any():
                all_missing_columns.append('KKS')
                missing_counts.append(((task_not_xx) & (kks_new_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (kks_new_na)].index + 3).tolist())

            if ((task_not_xx) & (equip_new_na)).any():
                all_missing_columns.append('EQUIPMENT')
                missing_counts.append(((task_not_xx) & (equip_new_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (equip_new_na)].index + 3).tolist())

            if ((task_not_xx) & (task_order_new_na)).any():
                all_missing_columns.append('TASK_ORDER')
                missing_counts.append(((task_not_xx) & (task_order_new_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (task_order_new_na)].index + 3).tolist())

            if ((task_not_xx) & (task_new_na)).any():
                all_missing_columns.append('TASK')
                missing_counts.append(((task_not_xx) & (task_new_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (task_new_na)].index + 3).tolist())

            if ((task_not_xx) & (start_date_na)).any():
                all_missing_columns.append('START_DATE')
                missing_counts.append(((task_not_xx) & (start_date_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (start_date_na)].index + 3).tolist())

            if ((task_not_xx) & (finish_date_na)).any():
                all_missing_columns.append('FINISH_DATE')
                missing_counts.append(((task_not_xx) & (finish_date_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (finish_date_na)].index + 3).tolist())

            if ((task_not_xx) & (ptw_na)).any():
                all_missing_columns.append('PTW')
                missing_counts.append(((task_not_xx) & (ptw_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (ptw_na)].index + 3).tolist())

            if ((task_not_xx) & (response_new_na)).any():
                all_missing_columns.append('RESPONSE')
                missing_counts.append(((task_not_xx) & (response_new_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (response_new_na)].index + 3).tolist())

            if ((task_not_xx) & (response_craft_na)).any():
                all_missing_columns.append('RESPONSE_CRAFT')
                missing_counts.append(((task_not_xx) & (response_craft_na)).sum())
                missing_indices.append((df_original_copy[(task_not_xx) & (response_craft_na)].index + 3).tolist())

            if (task_no_skill_rate).any():
                all_missing_columns.append('SKILL RATE')
                missing_counts.append(task_no_skill_rate.sum())
                missing_indices.append((df_original_copy[task_no_skill_rate].index + 3).tolist())

            if (task_no_type).any():
                all_missing_columns.append('TYPE')
                missing_counts.append(task_no_type.sum())
                missing_indices.append((df_original_copy[task_no_type].index + 3).tolist())

            # ตรวจสอบเงื่อนไขแต่ละคอลัมน์และเก็บชื่อคอลัมน์ที่ข้อมูลไม่ถูกต้อง
            invalid_columns = []
            invalid_counts = []
            invalid_indices = []

            if ((~cond_duration | non_negative) & task_order_not_xx).any():
                invalid_columns.append('DURATION_(HR.)')
                invalid_counts.append(((~cond_duration | non_negative) & task_order_not_xx).sum())
                invalid_indices.append((df_original[(~cond_duration | non_negative) & task_order_not_xx].index + 3).tolist())

            if (~task_order_valid & (df_original_check['TASK_ORDER'] != -1)).any():
                invalid_columns.append('TASK_ORDER')
                invalid_counts.append((~task_order_valid & (df_original_check['TASK_ORDER'] != -1)).sum())
                invalid_indices.append((df_original_check[(~task_order_valid & (df_original_check['TASK_ORDER'] != -1))].index + 3).tolist())

            if (cond_start_date_new).any():
                invalid_columns.append('START_DATE')
                invalid_counts.append(cond_start_date_new.sum())
                invalid_indices.append((df_original[cond_start_date_new].index + 3).tolist())

            if (cond_finish_date_new).any():
                invalid_columns.append('FINISH_DATE')
                invalid_counts.append(cond_finish_date_new.sum())
                invalid_indices.append((df_original[cond_finish_date_new].index + 3).tolist())

            if ((task_order) & (invalid_skill_rate)).any():
                invalid_columns.append('SKILL RATE')
                invalid_counts.append(((task_order) & (invalid_skill_rate)).sum())
                invalid_indices.append((df_original[(task_order) & (invalid_skill_rate)].index + 3).tolist())

            if ((task_type) & (~valid_type)).any():
                invalid_columns.append('TYPE')
                invalid_counts.append(((task_type) & (~valid_type)).sum())
                invalid_indices.append((df_original[(task_type) & (~valid_type)].index + 3).tolist())

            missing_messages = []
            invalid_messages = []
            error_messages= []

            # ตรวจสอบข้อมูลขาดหาย
            if all_missing_columns:
                missing_messages.extend(
                    [f"{col}: {count} รายการ {indices}"
                    for col, count, indices in zip(all_missing_columns, missing_counts, missing_indices)]
                )
                logger.error(f"Missing data detected in columns: {missing_messages}", exc_info=True)

            # ตรวจสอบข้อมูลที่ไม่ถูกต้อง
            if invalid_columns:
                invalid_messages.extend(
                    [f"{col}: {count} รายการ {indices}"
                    for col, count, indices in zip(invalid_columns, invalid_counts, invalid_indices)]
                )
                logger.error(f"Invalid data detected in columns: {invalid_messages}", exc_info=True)
            

            if missing_messages or invalid_messages:
                error_messages = "\n".join(missing_messages + invalid_messages)
                return render(request, 'maximo_app/upload_form.html', {
                    'form': form,
                    'missing_messages': missing_messages,
                    'invalid_messages': invalid_messages,
                    'error_messages': error_messages,
                    'schedule_filename': schedule_filename,
                    'location_filename': location_filename,
                    'selected_order': selected_order,
                })
            #! END RECHECK


            #! Creat JOB PLAN TASK
            df_original_newcol['UNIT'] = df_original_newcol['KKS_NEW'].str[0:3]
            df_original_filter = df_original_newcol[df_original_newcol['TASK_ORDER_NEW']!='xx'].copy()
            df_original_filter['TASK_ORDER_NEW'] = df_original_filter['TASK_ORDER_NEW'].astype('int32')

            cond1 = (df_original_filter['TASK_ORDER_NEW']==10) & (df_original_filter['TASK_ORDER_NEW'].shift(1)!=10)
            cond2 = ((df_original_filter['TASK_ORDER_NEW']==10) & (df_original_filter['TASK_ORDER_NEW'].shift(1)==10) & (df_original_filter['DURATION_(HR.)']!=-1))
            xx = cond1|cond2
            df_original_filter['GROUP_LEVEL_1'] = xx.cumsum().rename('GROUP_LEVEL_1')
            # df_original_filter['GROUP_LEVEL_2'] = df_original_filter['GROUP_LEVEL_1'].astype('str')+'-'+df_original_filter['KKS_NEW']
            df_original_filter['GROUP_LEVEL_2'] = (df_original_filter['GROUP_LEVEL_1'].astype(int) * 10).astype(str).str.zfill(5) + '-' + df_original_filter['KKS_NEW'].astype(str)
            df_original_filter['GROUP_LEVEL_2'] = df_original_filter['GROUP_LEVEL_2'].astype(str)
            df_original_filter['TYPE'] = df_original_filter['TYPE'].astype(str)
            df_original_filter['UNIT'] = df_original_filter['UNIT'].astype(str)
            df_original_filter['GROUP_LEVEL_3'] = np.where(
                worktype == 'APAO', 
                df_original_filter['GROUP_LEVEL_2'] + '-' + df_original_filter['TYPE'] + '-' + df_original_filter['UNIT'] + '-' + 'AD', 
                df_original_filter['GROUP_LEVEL_2'] + '-' + df_original_filter['TYPE'] + '-' + df_original_filter['UNIT'])
            common_indices = df_original_copy.index.intersection(df_original_filter.index)
            df_original_copy.loc[common_indices, ['GROUP_LEVEL_1', 'GROUP_LEVEL_2','GROUP_LEVEL_3']] = df_original_filter.loc[common_indices, ['GROUP_LEVEL_1', 'GROUP_LEVEL_2','GROUP_LEVEL_3']]
            common_indices1 = df_original.index.intersection(df_original_filter.index)
            df_original.loc[common_indices1, ['GROUP_LEVEL_1', 'GROUP_LEVEL_2','GROUP_LEVEL_3']] = df_original_filter.loc[common_indices1, ['GROUP_LEVEL_1', 'GROUP_LEVEL_2','GROUP_LEVEL_3']]

            data = df_original_filter.groupby(['GROUP_LEVEL_1','GROUP_LEVEL_3',
                                            'KKS_NEW','EQUIPMENT_NEW']).size()
            df_original_filter_group = pd.DataFrame(data).reset_index()
            df_original_filter_group = df_original_filter_group.drop(columns = [0])

            dict_jp_master = {}
            lst = ['GROUP_LEVEL_1','GROUP_LEVEL_3','KKS_NEW','EQUIPMENT_NEW','DURATION_TOTAL','TASK_ORDER_NEW', 'DURATION_(HR.)',
                'START_DATE','FINISH_DATE','TASK_NEW','RESPONSE_CRAFT', 'ประเภทของ_PERMIT_TO_WORK']
            
            for group_1 in df_original_filter_group['GROUP_LEVEL_1'].value_counts().sort_index().index:
                cond1 = (df_original_filter['GROUP_LEVEL_1']==group_1)
                work_group = df_original_filter[cond1].iloc[:,0:].reset_index(drop=True) # separate each group
                cond2 = work_group['DURATION_(HR.)']!=-1 # to eliminate redundant task
                df_new = work_group[cond2].copy()
                df_new['DURATION_TOTAL'] = df_new['DURATION_(HR.)'].sum()
                df_new = df_new.replace(-1,np.nan)
                df_new = df_new.ffill()
                df_new = df_new.fillna(-1)
                dict_jp_master[group_1] = df_new[lst]

            df_jop_plan = pd.DataFrame()
            for group1,group3,eq in zip(df_original_filter_group['GROUP_LEVEL_1'],
                                        df_original_filter_group['GROUP_LEVEL_3'],
                                        df_original_filter_group['EQUIPMENT_NEW']):
                
                df = dict_jp_master[group1].copy()
                lst = ['DURATION_TOTAL','TASK_ORDER_NEW','DURATION_(HR.)','START_DATE','FINISH_DATE','TASK_NEW']
                df_new = df[lst].copy()
                df_new.loc[:, 'JOB_NUM'] = [group3] * len(df_new)
                df_new.loc[:, 'EQUIPMENT'] = [eq] * len(df_new)
                lst2 = ['JOB_NUM','EQUIPMENT','DURATION_TOTAL','TASK_ORDER_NEW','DURATION_(HR.)','START_DATE','FINISH_DATE','TASK_NEW']
                df_jop_plan = pd.concat([df_jop_plan,df_new[lst2]])
            df_jop_plan = df_jop_plan.reset_index(drop=True)
            df_jop_plan['GROUP_NUM'] = df_jop_plan.groupby('JOB_NUM').ngroup()+1
            df1 = df_jop_plan['GROUP_NUM'].drop_duplicates()
            df1 = df1.reset_index(drop=True)
            df2 = pd.Series(np.arange(1,len(df1)+1),name = 'MAP_GROUP')
            df3 = pd.concat([df2,df1],axis=1)
            df3 = df3.set_index('GROUP_NUM')
            dict_new = df3.to_dict()['MAP_GROUP']
            df_jop_plan['GROUP']=df_jop_plan['GROUP_NUM'].map(dict_new)

            df_jop_plan['ORGID']= orgid
            df_jop_plan['SITEID']= siteid
            df_jop_plan['PLUSCREVNUM']= pluscrevum
            df_jop_plan['STATUS']= status
            df_jop_plan['PLUSCJPREVNUM']= pluscjprevnum

            new_columns= ['JOB_NUM', 'ORGID','SITEID','PLUSCREVNUM','STATUS',
                                'EQUIPMENT', 'DURATION_TOTAL', 'TASK_ORDER_NEW','PLUSCJPREVNUM',
                                'DURATION_(HR.)', 'TASK_NEW','START_DATE', 'FINISH_DATE',
                                'GROUP']
            df_jop_plan_master = df_jop_plan[new_columns].copy()
            df_jop_plan_master['MOD'] = df_jop_plan_master['GROUP']%2
            df_jop_plan_master['JOB_NUM'] = 'JP'+'-'+df_jop_plan_master['JOB_NUM']

            df_jop_plan_master['COMMENT'] = ''
            job_plan_cond1 = df_jop_plan_master['JOB_NUM'].astype(str).str.len()>30
            update_comment(df_jop_plan_master, job_plan_cond1, 'COMMENT', 'JPNUM มีความยาวมากกว่า 30 ตัวอักษร')


            job_plan_cond2 = df_jop_plan_master['EQUIPMENT'].astype(str).str.len()>100
            update_comment(df_jop_plan_master, job_plan_cond2, 'COMMENT', 'DESCRIPTION มีความยาวมากกว่า 100 ตัวอักษร')

            job_plan_cond3 = df_jop_plan_master['TASK_NEW'].astype(str).str.len()>100
            update_comment(df_jop_plan_master, job_plan_cond3, 'COMMENT', 'JOBTASK มีความยาวมากกว่า 100 ตัวอักษร')
            
            # df_jop_plan_master.to_excel(job_plan_task_path,index=False)


            #! JOB PLAN Labor
            df_jop_plan_master_labor = df_jop_plan_master.copy()
            df_jop_plan_master_labor['GROUP_LEVEL_1'] = df_jop_plan_master_labor['JOB_NUM'].str.extract(r'JP-(\d+)-')
            df_jop_plan_master_labor['GROUP_LEVEL_1'] = df_jop_plan_master_labor['GROUP_LEVEL_1'].astype('int32')
            df_jop_plan_master_labor['GROUP_LEVEL_1'] = df_jop_plan_master_labor['GROUP_LEVEL_1'] / 10

            df_group_level_1 = pd.DataFrame()
            for group_1 in df_original_filter_group['GROUP_LEVEL_1'].value_counts().sort_index().index:## get group 1
                cond1 = (df_original_filter['GROUP_LEVEL_1']==group_1)
                work_group = df_original_filter[cond1].iloc[:,0:].reset_index(drop=True) # separate each group
                cond2 = work_group['DURATION_(HR.)']!=-1 # to eliminate redundant task
                df_new = work_group[cond2].copy()
                df_new['DURATION_TOTAL']=df_new['DURATION_(HR.)'].sum()
                df_new = df_new.replace(-1,np.nan)
                df_new = df_new.ffill()
                df_new = df_new.fillna(-1)
                lst = ['RESPONSE_CRAFT','SUPERVISOR','FOREMAN','SKILL','DURATION_TOTAL','GROUP_LEVEL_1']
                df_group_level_1 = pd.concat([df_group_level_1,df_new[lst]])

            df_group_level_1 = df_group_level_1.reset_index(drop=True)

            df_labor = pd.DataFrame()
            for group_level_1 in df_jop_plan_master_labor['GROUP_LEVEL_1'].value_counts().sort_index().index:
                df1 = df_jop_plan_master_labor[df_jop_plan_master_labor['GROUP_LEVEL_1']==group_level_1][['JOB_NUM','DURATION_TOTAL','GROUP_LEVEL_1']]
                df2 = df_group_level_1[df_group_level_1['GROUP_LEVEL_1']==group_level_1]
                df1 = df1.reset_index(drop=True)
                df2 = df2.reset_index(drop=True)
                dfx = pd.concat([df1,df2],axis=1)
                _, i = np.unique(dfx.columns, return_index=True) ## Eliminate redundant columns
                i.sort()
                dfxx = dfx.iloc[:, i].copy()
                dfxx = dfxx.rename(columns={"SUPERVISOR": "L21NOM", "FOREMAN": "L22NOM", "SKILL":"L23NOM"})
                df_labor = pd.concat([df_labor,dfxx.iloc[0:1,0:]])

            df_labor = df_labor.reset_index(drop=True)
            df_labor.replace(-1,np.nan, inplace=True)
            df_melt = pd.melt(df_labor, id_vars=['JOB_NUM','DURATION_TOTAL','GROUP_LEVEL_1','RESPONSE_CRAFT'], value_vars=['L21NOM', 'L22NOM','L23NOM'])
            df_labor = df_melt.sort_values(by = ['GROUP_LEVEL_1','variable']).dropna()
            df_labor = df_labor.reset_index(drop=True)
            lst = ['JOB_NUM', 'RESPONSE_CRAFT', 'variable','DURATION_TOTAL', 'value']
            df_labor_new = df_labor[lst].copy()
            df_labor_new.columns = ['JOB_NUM', 'CRAFT', 'SKILLLEVEL', 'LABORHRS', 'QUANTITY']
            df_labor_new['GROUP_NUM'] = df_labor_new.groupby('JOB_NUM').ngroup()+1
            df1 = df_labor_new['GROUP_NUM'].drop_duplicates()
            df1 = df1.reset_index(drop=True)
            df2 = pd.Series(np.arange(1,len(df1)+1),name = 'MAP_GROUP')
            df3 = pd.concat([df2,df1],axis=1)
            df3 = df3.set_index('GROUP_NUM')
            dict_new = df3.to_dict()['MAP_GROUP']
            
            df_labor_new['GROUP']=df_labor_new['GROUP_NUM'].map(dict_new)
            df_labor_new['MOD'] = df_labor_new['GROUP']%2
            df_labor_new['ORGID'] = orgid
            df_labor_new['SITEID'] = siteid
            df_labor_new['PLUSCREVNUM']= pluscrevum
            df_labor_new['STATUS']= status
            df_labor_new['JPTASK']= ''
            df_labor_new['COMMENT'] = ''
            
            job_plan_cond1 = df_labor_new['JOB_NUM'].astype(str).str.len()>30
            update_comment(df_labor_new, job_plan_cond1, 'COMMENT', 'JPNUM มีความยาวมากกว่า 30 ตัวอักษร')
            
            lst_labor1 = ['JOB_NUM','ORGID','SITEID','PLUSCREVNUM', 
                'STATUS','JPTASK', 'CRAFT','SKILLLEVEL',
                'LABORHRS','QUANTITY','GROUP','MOD', 'COMMENT']
            
            # df_labor_new[lst_labor1].to_excel(job_plan_labor_path, index=False)


            #! Create PM PLAN
            # 'MAIN_SYSTEM','MAIN_SYSTEM_DESC','EGCRAFT','PTW'
            # logger.info(f"Grouping Options1 : {selected_order}")
            
            group_columns = selected_order
            
            pm_master_df = create_pm_plan(request, df_original_filter, siteid,
                                            first_plant, worktype, egmntacttype,
                                            wostatus, egprojectid, egwbs, frequency)
            
            if 'no_arrange' in group_columns:
                pm_master_df['MOD'] = 1
                pm_master_df['GROUP'] = ''
                df_pm_plan3 = pm_master_df.copy()
            else:
                if 'SYSTEM' in group_columns:
                    index = group_columns.index('SYSTEM')
                    group_columns[index:index+1] = ['MAIN_SYSTEM', 'MAIN_SYSTEM_DESC']
                
                group_base = ['UNIT', 'TYPE']
                all_group_columns = group_columns + group_base
                
                df_yy = pm_master_df.groupby(all_group_columns).size()
                level_to_sort = [all_group_columns.index('TYPE'), all_group_columns.index('UNIT')]
                df_yy = df_yy.sort_index(level=level_to_sort, ascending=False)
                df_yyy = df_yy.reset_index(name='count')
                df_yyy_cont_more = df_yyy[df_yyy['count']>1].reset_index(drop=True).copy()
                df_yyy_cont_less = df_yyy[df_yyy['count']<=1].reset_index(drop=True).copy()
                
                if not df_yyy_cont_more.empty:
                    df_pm_plan3 = group_pm_plan(request, pm_master_df, all_group_columns,
                                                    df_yyy_cont_more, df_yyy_cont_less, siteid,
                                                    group_columns, first_plant, location, worktype,
                                                    egmntacttype, wostatus, egprojectid, egwbs, frequency)
                else:
                    pm_master_df['MOD'] = 1
                    pm_master_df['GROUP'] = ''
                    df_pm_plan3 = pm_master_df.copy()

            df_pm_plan3['PARENTCHGSSTATUS']=''
            df_pm_plan3['WOSEQUENCE']=''

            lst_new = ['PMNUM', 'SITEID', 'DESCRIPTION', 'STATUS','LOCATION','WORKTYPE',
                    'EGMNTACTTYPE','WOSTATUS','EGPROJECTID','EGWBS', 'FREQUENCY', 'FREQUNIT',
                    'JPNUM', 'ROUTE', 'NEXTDATE', 'EGCRAFT', 'RESPONSED BY', 'PTW', 'PARENTCHGSSTATUS',
                    'LEADTIME', 'TARGSTARTTIME', 'PARENT', 'PMCOUNTER', 'WOSEQUENCE',
                    'MOD','MAIN_SYSTEM','MAIN_SYSTEM_DESC','UNIT','TYPE', 'SUB_SYSTEM', 'EQUIPMENT', 'SUB_SYSTEM_DESC',
                    'KKS_NEW_DESC', 'GROUP', 'FINISH_DATE', 'FINISH TIME']

            df_pm_plan3_master = df_pm_plan3[lst_new].reset_index(drop=True)
            index_parent = df_pm_plan3_master[(df_pm_plan3_master['PARENT']=='')].index
            index_child = df_pm_plan3_master[(df_pm_plan3_master['PARENT']!='')].index

            df_pm_plan3_master.loc[index_parent,'PARENTCHGSSTATUS']=0
            df_pm_plan3_master.loc[index_child,'PARENTCHGSSTATUS']=1
            df_pm_plan3_master.loc[:,'PMCOUNTER']=0

            index_parent2 = df_pm_plan3_master[df_pm_plan3_master['PARENTCHGSSTATUS'] == 0].index
            lst_parent_group = [i for i in range(1,len(index_parent2)+1)]
            df_pm_plan3_master.loc[index_parent2,'WOSEQUENCE'] = lst_parent_group

            for group in df_pm_plan3_master['PARENT'].value_counts().index:
                if group!='':
                    index_child = df_pm_plan3_master[df_pm_plan3_master['PARENT']==group].index
                    run_number = [i for i in range(1,len(index_child)+1)]
                    df_pm_plan3_master.loc[index_child,'WOSEQUENCE'] = run_number

            df_pm_plan3_master['NEXTDATE'] = df_pm_plan3_master['NEXTDATE'].astype('str')
            df_pm_plan3_master['FINISH_DATE'] = df_pm_plan3_master['FINISH_DATE'].astype('str')
            df_pm_plan3_master['TARGSTARTTIME'] = df_pm_plan3_master['TARGSTARTTIME'].astype('str')
            df_pm_plan3_master['FINISH TIME'] = df_pm_plan3_master['FINISH TIME'].astype('str')

            df_pm_plan3_master['COMMENT'] = ''
            pm_plan_cond0 = df_pm_plan3_master['PMNUM'].astype(str).str.len()>30
            update_comment(df_pm_plan3_master, pm_plan_cond0, 'COMMENT', 'PMNUM มีความยาวมากกว่า 30 ตัวอักษร')

            # df_pm_plan3_master.to_excel(pm_plan_path, index=False)


            #! TYPE 2
            class PMNumGenerator:
                def __init__(self):
                    self.primary_counter = 0
                    self.secondary_counter = 0
                    self.sub_counter = 0
                    self.pmnum0 = None
                
                def create_pmnum(self, row):
                    self.primary_counter += 10
                    try:
                        if row['PARENTCHGSSTATUS'] == 0 and (row['JPNUM'] == '' or pd.isna(row['JPNUM'])):
                            self.secondary_counter += 1
                            self.sub_counter = 0
                            if worktype == 'APAO':
                                self.pmnum0 = f"MI-{location}-AD-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}-{str(self.sub_counter).zfill(2)}"
                            else:
                                self.pmnum0 = f"MI-{location}-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}-{str(self.sub_counter).zfill(2)}"
                            pmnum1 = self.pmnum0
                            jpnum1 = ''
                            parent1 = ''
                        
                        elif row['PARENTCHGSSTATUS'] == 1 and pd.notna(row['JPNUM']):
                            self.sub_counter += 1
                            if worktype == 'APAO':
                                pmnum1 = f"MI-{location}-AD-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}-{str(self.sub_counter).zfill(2)}"
                            else:
                                pmnum1 = f"MI-{location}-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}-{str(self.sub_counter).zfill(2)}"
                            jpnum1 = f"JP-{pmnum1}"
                            parent1 = self.pmnum0
                        
                        else:
                            self.secondary_counter += 1
                            if worktype == 'APAO':
                                pmnum1 = f"MI-{location}-AD-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}"
                            else:
                                pmnum1 = f"MI-{location}-{str(self.primary_counter).zfill(5)}-{row['TYPE']}{str(self.secondary_counter).zfill(2)}"
                            jpnum1 = f"JP-{pmnum1}"
                            parent1 = ''
                    
                    except KeyError as e:
                        logger.error(f"Column not found: {e}")
                        error_message = (
                            f"<div class='error-container'>"
                            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบข้อมูลที่จำเป็นสำหรับการดำเนินการ<br>"
                            f"<ul class='error-details'>"
                            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                            f"<li>คอลัมน์ที่ขาดหายไป: {str(e)}</li>"
                            f"</ul>"
                            f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าชื่อคอลัมน์ถูกต้องและครบถ้วน  หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                            f"</div>"
                        )
                        messages.error(request, error_message)
                        return redirect('index')
                    except Exception as e:
                        logger.error(f"Error processing row: {e}")
                        error_message = (
                            f"<div class='error-container'>"
                            f"<strong class='error-title'>พบปัญหา:</strong> เกิดข้อผิดพลาดระหว่างการประมวลผลข้อมูล<br>"
                            f"<ul class='error-details'>"
                            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                            f"<li>{str(e)}</li>"
                            f"</ul>"
                            f"<p class='error-note'>คำแนะนำ: กรุณาลองดำเนินการอีกครั้ง หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                            f"</div>"
                        )
                        messages.error(request, error_message)
                        return redirect('index')
                    
                    return pd.Series([self.pmnum0, pmnum1, jpnum1, parent1])

            generator = PMNumGenerator()
            df_pm_plan3_master[['PMNUM1', 'PMNUM1', 'JPNUM1', 'PARENT1']] = df_pm_plan3_master.apply(generator.create_pmnum, axis=1)
            
            jpnum_map = df_pm_plan3_master.set_index('JPNUM')['JPNUM1'].to_dict()
            df_jop_plan_master['JOB_NUM1'] = df_jop_plan_master['JOB_NUM'].map(jpnum_map)
            df_labor_new['JOB_NUM1'] = df_labor_new['JOB_NUM'].map(jpnum_map)

            df_pm_plan3_master['COMMENT1'] = ''
            pm_plan_cond1 = df_pm_plan3_master['PMNUM1'].astype(str).str.len()>30
            update_comment(df_pm_plan3_master, pm_plan_cond1, 'COMMENT1', 'PMNUM มีความยาวมากกว่า 30 ตัวอักษร')

            df_jop_plan_master['COMMENT1'] = ''
            job_plan_cond1 = df_jop_plan_master['JOB_NUM1'].astype(str).str.len()>30
            update_comment(df_jop_plan_master, job_plan_cond1, 'COMMENT1', 'JPNUM มีความยาวมากกว่า 30 ตัวอักษร')

            df_labor_new['COMMENT1'] = ''
            job_plan_cond2 = df_labor_new['JOB_NUM1'].astype(str).str.len()>30
            update_comment(df_labor_new, job_plan_cond2, 'COMMENT1', 'JPNUM มีความยาวมากกว่า 30 ตัวอักษร')
            
            pm_plan_column1 = ['PMNUM', 'SITEID', 'DESCRIPTION', 'STATUS', 'LOCATION', 'WORKTYPE',
                                    'EGMNTACTTYPE', 'WOSTATUS', 'EGPROJECTID', 'EGWBS', 'FREQUENCY',
                                    'FREQUNIT', 'JPNUM', 'ROUTE', 'NEXTDATE', 'EGCRAFT', 'RESPONSED BY',
                                    'PTW', 'PARENTCHGSSTATUS', 'LEADTIME', 'TARGSTARTTIME', 'PARENT',
                                    'PMCOUNTER', 'WOSEQUENCE', 'MOD', 'MAIN_SYSTEM', 'MAIN_SYSTEM_DESC',
                                    'UNIT', 'TYPE', 'SUB_SYSTEM', 'EQUIPMENT', 'SUB_SYSTEM_DESC',
                                    'KKS_NEW_DESC', 'GROUP', 'FINISH_DATE', 'FINISH TIME', 'COMMENT']
            replace_dict = {
                'PMNUM': 'PMNUM1',
                'JPNUM': 'JPNUM1',
                'PARENT': 'PARENT1',
                'COMMENT': 'COMMENT1'
            }
            pm_plan_column2 = [replace_columns(col, replace_dict) for col in pm_plan_column1]

            jop_plan_column1 = ['JOB_NUM', 'ORGID', 'SITEID', 'PLUSCREVNUM', 'STATUS', 'EQUIPMENT', 
                                    'DURATION_TOTAL', 'TASK_ORDER_NEW', 'PLUSCJPREVNUM', 'DURATION_(HR.)', 
                                    'TASK_NEW', 'START_DATE', 'FINISH_DATE', 'GROUP', 'MOD', 'COMMENT']
            replace_dict = {
                'JOB_NUM': 'JOB_NUM1',
                'COMMENT': 'COMMENT1',
            }
            jop_plan_column2 = [replace_columns(col, replace_dict) for col in jop_plan_column1]
            lst_labor2 = [replace_columns(col, replace_dict) for col in lst_labor1]
            
            #! Create WORKORDER
            df_workorder1 = create_workorder(request, df_pm_plan3_master, 'PMNUM', orgid)
            df_workorder2 = create_workorder(request, df_pm_plan3_master, 'PMNUM1', orgid)
            
            #! Create Template-MxLoader-JP-PMPlan
            # Define the color and border style
            color = "666666"
            thin_border = Border(
                left=Side(style='thin', color=color),
                right=Side(style='thin', color=color),
                top=Side(style='thin', color=color),
                bottom=Side(style='thin', color=color)
            )

            # Define the background color fill
            fill_color = PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Define file paths
            mxloader_template_v842_path = os.path.join(settings.STATIC_ROOT, 'excel', 'Template MxLoader JB-PM Plan.xlsm')
            mxloader_template_v810_path = os.path.join(settings.STATIC_ROOT, 'excel', 'Template MxLoader JB-PM Plan v8.1.0.xlsm')
            file_template_xlsx = os.path.join(temp_dir, f"{uuid.uuid4()}_Template MxLoader JB-PM Plan_temp.xlsx")
            mxloader_template_output_v842_path = os.path.join(temp_dir, f"{uuid.uuid4()}_Template MxLoader JB-PM Plan.xlsm")
            mxloader_template_output_v810_path = os.path.join(temp_dir, f"{uuid.uuid4()}_Template MxLoader JB-PM Plan v.8.1.0.xlsm")
            
            sheet_jp_labor = 'JPPLAN-LABOR'
            sheet_jp_task = 'JPPLAN-TASK'
            sheet_pm = 'PMPlan'
            sheet_wo = 'WO'
            start_row = 2

            if not os.path.exists(mxloader_template_v842_path):
                raise FileNotFoundError(f"The file {mxloader_template_v842_path} does not exist.")

            shutil.copyfile(mxloader_template_v842_path, file_template_xlsx)

            try:
                basic_sheet_names = [sheet_jp_labor, sheet_jp_task, sheet_pm, sheet_wo]
                for i in range(1, 2):
                    for sheet_name in basic_sheet_names:
                        copy_worksheet(request, file_template_xlsx, sheet_name, i)
                
                with pd.ExcelWriter(file_template_xlsx, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # ชุดที่ 1
                    write_dataframes_to_excel(
                        request,
                        writer, 
                        df_labor_new, lst_labor1, 
                        df_jop_plan_master, jop_plan_column1, 
                        df_pm_plan3_master, pm_plan_column1, 
                        df_workorder1, 
                        sheet_jp_labor, sheet_jp_task, sheet_pm, sheet_wo, 
                        start_row, start_offset=5
                    )
                    
                    # ชุดที่ 2
                    write_dataframes_to_excel(
                        request,
                        writer, 
                        df_labor_new, lst_labor2, 
                        df_jop_plan_master, jop_plan_column2, 
                        df_pm_plan3_master, pm_plan_column2, 
                        df_workorder2, 
                        f"{sheet_jp_labor}-1", f"{sheet_jp_task}-1", f"{sheet_pm}-1", f"{sheet_wo}-1", 
                        start_row, start_offset=5
                    )

                openpyxl_book = load_workbook(file_template_xlsx)
                
                for sheet_name in basic_sheet_names:
                    sheet = openpyxl_book[sheet_name]
                    decorate_sheet(sheet, sheet_name, thin_border, fill_color, yellow_fill, start_row, df_jop_plan_master, df_pm_plan3_master)

                for sheet_name in basic_sheet_names:
                    sheet_name_with_suffix = f"{sheet_name}-1"
                    sheet = openpyxl_book[sheet_name_with_suffix]
                    decorate_sheet(sheet, sheet_name_with_suffix, thin_border, fill_color, yellow_fill, start_row, df_jop_plan_master, df_pm_plan3_master)
                
                openpyxl_book.save(file_template_xlsx)
                
                copy_sheets_to_macro_file(
                    request, file_template_xlsx, mxloader_template_v842_path, mxloader_template_output_v842_path, basic_sheet_names, location
                    )
                
                copy_sheets_to_macro_file(
                    request, file_template_xlsx, mxloader_template_v810_path, mxloader_template_output_v810_path, basic_sheet_names, location
                    )

            except Exception as e:
                logger.error(f"An error occurred: {str(e)}", exc_info=True)
                error_message = (
                    f"<div class='error-container'>"
                    f"<strong class='error-title'>พบปัญหา:</strong> เกิดข้อผิดพลาดระหว่างการดำเนินการกับไฟล์ข้อมูล<br>"
                    f"<ul class='error-details'>"
                    f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                    f"<li>{str(e)}</li>"
                    f"</ul>"
                    f"<p class='error-note'>กรุณาลองดำเนินการอีกครั้ง หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                    f"</div>"
                )
                messages.error(request, error_message)
                return redirect('index')
            
            finally:
                try:
                    openpyxl_book.close()
                except Exception as e:
                    logger.error(f"Failed to close the openpyxl workbook: {e}")
                    error_message = (
                        f"<div class='error-container'>"
                        f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถปิดไฟล์ Excel ได้<br>"
                        f"<ul class='error-details'>"
                        f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                        f"<li>{str(e)}</li>"
                        f"</ul>"
                        f"<p class='error-note'>คำแนะนำ: กรุณาลองดำเนินการอีกครั้ง หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                        f"</div>"
                    )
                    messages.error(request, error_message)
                    return redirect('index')
                
                try:
                    if os.path.exists(file_template_xlsx):
                        os.remove(file_template_xlsx)
                except FileNotFoundError:
                    logger.warning(f"Temporary file {file_template_xlsx} not found for deletion.")
                except Exception as e:
                    logger.error(f"Failed to delete temporary file: {e}")
                    error_message = (
                        f"<div class='error-container'>"
                        f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถลบไฟล์ชั่วคราวได้<br>"
                        f"<ul class='error-details'>"
                        f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                        f"<li>{str(e)}</li>"
                        f"</ul>"
                        f"<p class='error-note'>คำแนะนำ: กรุณาลองดำเนินการอีกครั้ง หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
                        f"</div>"
                    )
                    messages.error(request, error_message)
                    return redirect('index')
            
            # Save session
            request.session['first_plant'] = first_plant
            request.session['most_common_plant_unit'] = most_common_plant_unit
            
            request.session['download_link_template'] = {
                '8.4.2': mxloader_template_output_v842_path,
                '8.1.0': mxloader_template_output_v810_path,
            }
        
    else:
        keys_to_clear = [
            'schedule_filename', 'location_filename', 'temp_dir', 'schedule_path', 'location_path', 
            'first_plant', 'most_common_plant_unit',
            'download_link_comment', 
            'download_link_template', 
            'year', 'frequency', 'egmntacttype', 'egprojectid', 'egwbs', 'location', 'siteid', 
            'wostatus', 'wbs', 'wbs_desc', 'worktype', 'wostatus', 'grouping_text', 'child_site', 
        ]
        # 'download_link_job_plan_task', 'download_link_job_plan_labor', 'download_link_pm_plan'
        
        for key in keys_to_clear:
            request.session.pop(key, None)
        
        form = UploadFileForm()
    
    return render(request, 'maximo_app/upload_form.html', {
        'form': form,
        'missing_messages': missing_messages,
        'invalid_messages': invalid_messages,
        'error_messages': error_messages,
        'schedule_filename': schedule_filename,
        'location_filename': location_filename,
        'selected_order': selected_order,
    })

# ---------------------------------
# ฟังก์ชันการจัดการการดาวน์โหลด (Download Functions)
# ---------------------------------
def generic_download(request, session_key, original_file_name, content_type, file_path):
    if session_key is None:
        file_path = file_path
    else:
        file_path = request.session.get(session_key, None)
    if not file_path:
        logger.error(f"File path for session key '{session_key}' not specified.")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่มีไฟล์ที่ต้องการดาวน์โหลด<br>"
            f"<ul class='error-details'>"
            f"<li>ระบบไม่ได้รับข้อมูลที่อยู่ของไฟล์จาก session</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาอัปโหลดไฟล์อีกครั้ง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    
    full_file_path = os.path.abspath(file_path) if not os.path.isabs(file_path) else file_path
    
    if not os.path.exists(full_file_path):
        logger.error(f"File '{original_file_name}' not found at path '{full_file_path}'.")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ {original_file_name} ที่ต้องการดาวน์โหลด<br>"
            f"<ul class='error-details'>"
            f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
            f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาอัปโหลดไฟล์อีกครั้ง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    
    try:
        with open(full_file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{original_file_name}"'
            return response
    except Exception as e:
        logger.error(f"Failed to open file '{original_file_name}' for download: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถเปิดไฟล์ {original_file_name} ได้<br>"
            f"<ul class='error-details'>"
            f"<li>เกิดข้อผิดพลาดระหว่างการเปิดไฟล์</li>"
            f"<li>กรุณาตรวจสอบว่าไฟล์มีอยู่จริงและไม่เสียหาย</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาอัปโหลดไฟล์อีกครั้ง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_comment_file(request):
    return generic_download(
        request=request,
        session_key='download_link_comment',
        original_file_name='Comment.xlsx',
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_path=None
    )

# def download_job_plan_task_file(request):
#     return generic_download(
#         request=request,
#         session_key='download_link_job_plan_task',
#         original_file_name='Job_Plan_Task.xlsx',
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         file_path=None
#     )

# def download_job_plan_labor_file(request):
#     return generic_download(
#         request=request,
#         session_key='download_link_job_plan_labor',
#         original_file_name='Job_Plan_Labor.xlsx',
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         file_path=None
#     )

# def download_pm_plan_file(request):
#     return generic_download(
#         request=request,
#         session_key='download_link_pm_plan',
#         original_file_name='PM_Plan.xlsx',
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         file_path=None
#     )

def download_user_manual(request):
    try:
        file_path = os.path.join(settings.STATIC_ROOT, 'pdf', 'คู่มือการใช้งาน.pdf')
        
        if not os.path.exists(file_path):
            logger.error("User manual file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์คู่มือการใช้งานที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')
        
        # สร้าง FileResponse สำหรับส่งไฟล์ PDF ให้ผู้ใช้ดาวน์โหลด
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download user manual file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์คู่มือการใช้งานได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_template_file(request, version):
    location = request.session.get('location', 'BLANK')
    download_links = request.session.get('download_link_template', {})
    file_path = download_links.get(version) # ดึง file_path ตามเวอร์ชัน
    
    if not file_path or not isinstance(file_path, str):
        logger.error(f"No valid download link found for version {version}.")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Template MxLoader JB-PM Plan เวอร์ชัน {version} ที่ต้องการดาวน์โหลด<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
            f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้วหรือถูกลบออก</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาอัปโหลดไฟล์อีกครั้ง หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    
    if not os.path.exists(file_path):
        logger.error(f"File path does not exist: {file_path}")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Template MxLoader JB-PM Plan เวอร์ชัน {version} ที่ต้องการดาวน์โหลด<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>ไฟล์อาจถูกลบ หรือหมดอายุแล้ว</li>"
            f"<li>>อาจเกิดข้อผิดพลาดระหว่างการอัปโหลด หรือไฟล์ยังไม่ถูกสร้าง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาอัปโหลดไฟล์อีกครั้ง หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    
    # กำหนด session_key และ original_file_name ตามเวอร์ชัน
    if version == '8.4.2':
        original_file_name = f'Template MxLoader JB-PM Plan_{location}.xlsm'
    elif version == '8.1.0':
        original_file_name = f'Template MxLoader JB-PM Plan v8.1.0_{location}.xlsm'
    else:
        logger.error(f"Invalid version specified: {version}")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> เวอร์ชันที่ระบุ ({version}) ไม่ถูกต้อง<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เวอร์ชันที่ระบุไม่มีอยู่ในระบบ</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        return redirect('index')
    
    return generic_download(
        request=request,
        session_key=None,
        original_file_name=original_file_name,
        content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        file_path=file_path
    )

def download_original_template(request):
    try:
        file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'Template MxLoader JB-PM Plan.xlsm')
        
        if not os.path.exists(file_path):
            logger.error("Template MxLoader JB-PM Plan file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Template MxLoader JB-PM Plan ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')
        
        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download Template MxLoader JB-PM Plan file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ Template MxLoader JB-PM Plan ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_schedule(request):
    try:
        file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'Final Schedule.xlsx')
        
        if not os.path.exists(file_path):
            logger.error("Final Schedule file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Final Schedule ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')
        
        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download Final Schedule file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ Final Schedule ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_example_template(request):
    try:
        file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'TEMPLATE MXLOADERก JB-PM PLAN(EXAMPLE).zip')
        
        if not os.path.exists(file_path):
            logger.error("TEMPLATE MXLOADER JB-PM PLAN(EXAMPLE) zip file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ TEMPLATE MXLOADER JB-PM PLAN(EXAMPLE) ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')

        response = FileResponse(open(file_path, 'rb'), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download TEMPLATE MXLOADER JB-PM PLAN(EXAMPLE) zip file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ TEMPLATE MXLOADER JB-PM PLAN(EXAMPLE) ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_example_schedule(request):
    try:
        file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'FINAL SCHEDULE(EXAMPLE).zip')
        
        if not os.path.exists(file_path):
            logger.error("FINAL SCHEDULE(EXAMPLE) zip file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Final Schedule ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')
        
        response = FileResponse(open(file_path, 'rb'), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download FINAL SCHEDULE(EXAMPLE) zip file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ Final Schedule ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_user_schedule(request):
    try:
        file_path = request.session.get('schedule_path', None)
        file_name = request.session.get('schedule_filename', 'Final Schedule.xlsx')

        if not file_path or not os.path.exists(file_path):
            logger.error("Final Schedule file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Final Schedule ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')

        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download Final Schedule file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ Final Schedule ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def download_user_location(request):
    try:
        file_path = request.session.get('location_path', None)
        file_name = request.session.get('location_filename', 'Location.xlsx')

        if not file_path or not os.path.exists(file_path):
            logger.error("Location file not found at: %s", file_path)
            error_message = (
                f"<div class='error-container'>"
                f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Location ที่ต้องการดาวน์โหลด<br>"
                f"<ul class='error-details'>"
                f"<p class='error-description'>สาเหตุของปัญหา:</p>"
                f"<li>ระบบไม่สามารถหาไฟล์ที่ต้องการดาวน์โหลดได้</li>"
                f"<li>ไฟล์ที่ต้องการดาวน์โหลดได้หมดอายุแล้ว</li>"
                f"</ul>"
                f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าไฟล์มีอยู่จริง หรือติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
                f"</div>"
            )
            messages.error(request, error_message)
            return redirect('index')

        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response

    except Exception as e:
        logger.error(f"Failed to download Location file: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถดาวน์โหลดไฟล์ Location ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการดาวน์โหลดไฟล์ กรุณาลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')


# ---------------------------------
# ฟังก์ชันการกรองข้อมูล (Filter Functions)
# ---------------------------------
@require_GET
def filter_site(request):
    site_id = request.GET.get('site_id')
    
    if not site_id:
        return JsonResponse({'error': 'No site_id provided.'}, status=400)
    
    try:
        site_id = int(site_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid site_id format. Must be an integer.'}, status=400)
    
    try:
        site = Site.objects.get(id=site_id)
        # logger.info(f"Successfully retrieved Site with id {site_id}")
        child_sites = site.child_sites.values('id', 'site_id', 'site_name')
        
    except Site.DoesNotExist:
        logger.warning(f"Site with id {site_id} does not exist.")
        return JsonResponse({'error': 'Site not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_site: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)

    child_site_list = [{
        'id': child_site['id'],
        'site_id': child_site['site_id'],
        'site_name': child_site['site_name']
    } for child_site in child_sites]
    
    return JsonResponse({
        'site_name': site.site_name,
        'child_sites': child_site_list,
    }, status=200)

@require_GET
def filter_child_site(request):
    child_site_id = request.GET.get('child_site_id')

    if not child_site_id:
        return JsonResponse({'error': 'No child_site_id provided.'}, status=400)
    
    try:
        child_site = ChildSite.objects.get(id=child_site_id)
        # logger.info(f"Successfully retrieved ChildSite with id {child_site_id}")
    except ChildSite.DoesNotExist:
        logger.warning(f"ChildSite with id {child_site_id} does not exist.")
        return JsonResponse({'error': 'Child Site not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_child_site: {str(e)}")
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)

    return JsonResponse({
        'description': child_site.site_name
    }, status=200)

@require_GET
def filter_worktype(request):
    work_type_id = request.GET.get('work_type_id')
    
    if not work_type_id:
        return JsonResponse({'error': 'No work_type_id provided.'}, status=400)
    
    try:
        work_type_id = int(work_type_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid work_type_id format. Must be an integer.'}, status=400)
    
    try:
        work_type = WorkType.objects.get(id=work_type_id)
        # logger.info(f"Successfully retrieved WorkType with id {work_type_id}")
    except WorkType.DoesNotExist:
        logger.warning(f"WorkType with id {work_type_id} does not exist.")
        return JsonResponse({'error': 'WorkType not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_worktype: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)
    return JsonResponse({
        'description': work_type.description
    }, status=200)

@require_GET
def filter_plant_type(request):
    plant_type_id = request.GET.get('plant_type_id')
    
    if not plant_type_id:
        return JsonResponse({'error': 'No plant_type_id provided.'}, status=400)
    
    try:
        plant_type_id = int(plant_type_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid plant_type_id format. Must be an integer.'}, status=400)
    
    try:
        # ดึงข้อมูล PlantType ตาม id
        plant_type = PlantType.objects.get(id=plant_type_id)
        # logger.info(f"Successfully retrieved PlantType with id {plant_type_id}")
        
        # Get the ActTypes associated with the PlantType
        acttypes = plant_type.act_types.all()
        
        # Get the Sites associated with the PlantType
        sites = plant_type.sites.all()
        
        # Get the WorkType associated with the PlantType
        work_types = plant_type.work_types.all()
        
        # Get the Units associated with the PlantType
        units = plant_type.units.all()
    except PlantType.DoesNotExist:
        logger.warning(f"PlantType with id {plant_type_id} does not exist.")
        return JsonResponse({'error': 'PlantType not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_plant_type: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)
    
    # แปลงข้อมูลเป็น list เพื่อส่งกลับไปยัง frontend
    acttype_list = [
        {
            'id': acttype.id, 
            'acttype': acttype.acttype, 
            'description': acttype.description
        } for acttype in acttypes
    ]
    site_list = [
        {
            'id': site.id, 
            'site_id': site.site_id, 
            'site_name': site.site_name
        } for site in sites
    ]    
    work_type_list = [
        {
            'id': worktype.id, 
            'worktype': worktype.worktype,
            'description': worktype.description
        } for worktype in work_types
    ]
    unit_list = [
        {
            'id': unit.id, 
            'unit_code': unit.unit_code
        } for unit in units
    ]
    
    return JsonResponse({
        'acttypes': acttype_list,
        'sites': site_list,
        'plant_type_th': plant_type.plant_type_th,
        'work_types': work_type_list,
        'units' : unit_list,
    }, status=200)
    
@require_GET
def filter_acttype(request):
    acttype_id = request.GET.get('acttype_id')

    if not acttype_id:
        return JsonResponse({'error': 'No acttype_id provided.'}, status=400)

    try:
        acttype_id = int(acttype_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid acttype_id format. Must be an integer.'}, status=400)

    try:
        acttype = ActType.objects.get(id=acttype_id)
        # logger.info(f"Successfully retrieved ActType with id {acttype_id}")
    except ActType.DoesNotExist:
        logger.warning(f"ActType with id {acttype_id} does not exist.")
        return JsonResponse({'error': 'ActType not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_acttype: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)
    
    return JsonResponse({
        'description': acttype.description, 
        'code': acttype.code
    }, status=200)

@require_GET
def filter_wbs(request):
    wbs_id = request.GET.get('wbs_id')

    if not wbs_id:
        return JsonResponse({'error': 'No wbs_id provided.'}, status=400)

    try:
        wbs_id = int(wbs_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid wbs_id format. Must be an integer.'}, status=400)

    try:
        wbs = WBSCode.objects.get(id=wbs_id)
        # logger.info(f"Successfully retrieved WBSCode with id {wbs_id}")
    except WBSCode.DoesNotExist:
        logger.warning(f"WBSCode with id {wbs_id} does not exist.")
        return JsonResponse({'error': 'WBSCode not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_wbs: {e}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)
    
    return JsonResponse({
        'description': wbs.description
    }, status=200)

@require_GET
def filter_wostatus(request):
    wostatus_id = request.GET.get('wostatus_id')

    if not wostatus_id:
        return JsonResponse({'error': 'No wostatus_id provided.'}, status=400)

    try:
        wostatus_id = int(wostatus_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid wostatus_id format. Must be an integer.'}, status=400)

    try:
        wostatus = Status.objects.get(id=wostatus_id)
        # logger.info(f"Successfully retrieved Status with id {wostatus_id}")
    except Status.DoesNotExist:
        logger.warning(f"Status with id {wostatus_id} does not exist.")
        return JsonResponse({'error': 'Status not found.'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred in filter_wostatus: {e}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)
    
    return JsonResponse({
        'description': wostatus.description
    }, status=200)

# ---------------------------------
# ส่วนของ Custom Error Handlers
# ---------------------------------
def custom_404(request, exception):
    # ดึง stack trace ของเฟรมปัจจุบัน (ข้อมูลสำหรับ debug)
    stack_trace = ''.join(traceback.format_stack())
    
    error_details = (
        f"Page not found (404)\n"
        f"Request Method: {request.method}\n"
        f"Request URL: {request.build_absolute_uri()}\n\n"
        f"Stack trace (most recent call last):\n{stack_trace}\n\n"
        f"Django tried these URL patterns:\n{exception}"
    )
    logger.error(error_details)
    return render(request, 'errors/404.html', {'error_details': error_details}, status=404)

def custom_500(request):
    # ดึงรายละเอียดของ exception
    exc_type, exc_value, exc_traceback = sys.exc_info()

    error_trace = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))

    error_details = (
        f"Server Error (500)\n"
        f"Request Method: {request.method}\n"
        f"Request URL: {request.build_absolute_uri()}\n\n"
        f"{error_trace}"
    )
    return render(request, 'errors/500.html', {'error_details': error_details}, status=500)


# ---------------------------------
# ฟังก์ชันช่วยเหลือ (Helper Functions)
# ---------------------------------
def get_grouping_text(selected_order):
    if 'no_arrange' in selected_order:
        return 'ไม่มีการจัดกลุ่ม'
    return f"จัดกลุ่มตาม {', '.join(selected_order)}"

def update_comment(df, condition, column_name, message):
    comment_empty_or_na = (df[column_name] == '') | (df[column_name].isna())
    df.loc[condition & comment_empty_or_na, column_name] = message
    df.loc[condition & ~comment_empty_or_na, column_name] += f", {message}"

def replace_or_append_comment(df, condition, comment_col, message, replace_message=None):
    # กรณีที่ COMMENT ว่างหรือเป็น NaN ให้เขียนข้อความใหม่
    comment_empty_or_na = (df[comment_col] == '') | (df[comment_col].isna())
    df.loc[condition & comment_empty_or_na, comment_col] = message

    # กรณีที่ COMMENT มีข้อความแล้ว ให้เพิ่มหรือแทนที่ข้อความ
    def update_comment_text(existing_comment):
        if replace_message and replace_message in existing_comment.split(', '):
            return existing_comment.replace(replace_message, message)
        else:
            return f"{existing_comment}, {message}"

    df.loc[condition & ~comment_empty_or_na, comment_col] = df.loc[condition & ~comment_empty_or_na, comment_col].apply(update_comment_text)

def read_excel_with_error_handling(request, schedule_path, sheet_name=0, header=1, dtype_spec=None):
    try:
        df_original = pd.read_excel(schedule_path, sheet_name=sheet_name, header=header, dtype=dtype_spec)
        return df_original
    except ValueError as ve:
        # sheet_name ไม่มีข้อมูล
        logger.error(f"Invalid sheet name '{sheet_name}': {ve}")
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบชีท '{sheet_name}' ที่ระบุในไฟล์ Excel<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>ชื่อชีทที่ระบุอาจไม่ถูกต้องหรือไม่มีอยู่ในไฟล์</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบชื่อชีทและลองอีกครั้ง หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือ</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    except Exception as e:
        logger.error(f"Failed to read Excel file '{schedule_path}': {e}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถเปิดไฟล์ Excel ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการอ่านไฟล์ กรุณาตรวจสอบว่าไฟล์มีอยู่และสามารถเปิดได้</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)        
        return redirect('index')

def convert_duration(value):
    try:
        value = float(value)
        if value.is_integer():
            return int(value)
        else:
            return value
    except (ValueError, TypeError):
        return value

def is_date(value):
    try:
        # ตรวจสอบว่าค่าคือวันที่ในรูปแบบ "DD-MMM-YYYY"
        return bool(re.match(r"\d{1,2}-[A-Za-z]{3}-\d{4}", value))
    except TypeError:
        return False

def parse_dates(date_series):
    # แปลงวันที่ตามรูปแบบต่าง ๆ
    date_1 = pd.to_datetime(date_series, format='%d/%m/%Y', errors='coerce')
    date_2 = pd.to_datetime(date_series, format='%Y-%m-%d %H:%M:%S', errors='coerce')
    date_3 = pd.to_datetime(date_series, format='%d-%b-%Y', errors='coerce')
    date_4 = pd.to_datetime(date_series, format='%m/%d/%Y', errors='coerce')

    return date_1.fillna(date_2).fillna(date_3).fillna(date_4)

def clean_ptw_column(ptw_value):
    try:
        if not ptw_value or not isinstance(ptw_value, str):
            raise ValueError("ค่า PTW ต้องเป็นสตริงที่ไม่ว่าง")

        # แยกข้อความตาม '//'
        if '//' in ptw_value:
            parts = ptw_value.split('//')
            left_side = parts[0]
            right_side = ','.join([part.strip() for part in parts[1:]])
            left_side_list = [item.strip() for item in left_side.split(',')]
            right_side_set = set([item.strip() for item in right_side.split(',')])
            cleaned_right_side = [item for item in right_side_set if item not in left_side_list]
            cleaned_left_str = ', '.join(left_side_list)
            cleaned_right_str = ', '.join(cleaned_right_side)
            
            if cleaned_right_str:
                return f'{cleaned_left_str}//{cleaned_right_str}'
            else:
                return cleaned_left_str
        else:
            return ptw_value

    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return ptw_value

def create_pm_plan(request, df_original_filter, siteid,
                    first_plant, worktype, egmntacttype,
                    wostatus, egprojectid, egwbs, frequency):
    try:
        df_original_filter['ROUTE'] = df_original_filter['ROUTE'].replace(-1, np.nan)
        data_pm = df_original_filter.groupby(['GROUP_LEVEL_1','GROUP_LEVEL_3','KKS_NEW','MAIN_SYSTEM','SUB_SYSTEM',
                                    'EQUIPMENT','MAIN_SYSTEM_DESC','SUB_SYSTEM_DESC','KKS_NEW_DESC','EQUIPMENT_NEW',
                                    'UNIT','TYPE'
                                ]).size()
        df_original_filter_group_pm = pd.DataFrame(data_pm).reset_index()
        df_original_filter_group_pm = df_original_filter_group_pm.drop(columns = [0])
        pm_master_dict = {
            'PMNUM': [],
            'SITEID': [],
            'DESCRIPTION': [],
            'STATUS': [],
            'LOCATION': [],
            'ROUTE': [],
            'LEADTIME': [],
            'PMCOUNTER': [],
            'WORKTYPE': [],
            'EGMNTACTTYPE': [],
            'WOSTATUS': [],
            'EGCRAFT': [],
            'RESPONSED BY': [],
            'PTW': [],
            'LOTO': [],
            'EGPROJECTID': [],
            'EGWBS': [],
            'FREQUENCY': [],
            'FREQUNIT': [],
            'NEXTDATE': [],
            'TARGSTARTTIME': [],
            'FINISH_DATE': [],
            'FINISH TIME': [],
            'PARENT': [],
            'JPNUM': [],
            'MAIN_SYSTEM': [],
            'SUB_SYSTEM': [],
            'EQUIPMENT': [],
            'MAIN_SYSTEM_DESC': [],
            'SUB_SYSTEM_DESC': [],
            'KKS_NEW_DESC': [],
            'UNIT': [],
            'TYPE': [],
        }

        for _,row in  df_original_filter_group_pm.iterrows():
            #row_index
            ####PM Number
            df_group_temp = df_original_filter.copy()
            PMNUM = 'PO'+'-'+row['GROUP_LEVEL_3']
            pm_master_dict['PMNUM'].append(PMNUM)
            #####Site ID
            pm_master_dict['SITEID'] = siteid
            #####DESCRIPTION
            desc = row['EQUIPMENT_NEW']
            pm_master_dict['DESCRIPTION'].append(desc)
            #####STATUS
            pm_master_dict['STATUS'] = status
            #####LOCATION
            loc = first_plant+row['KKS_NEW']
            pm_master_dict['LOCATION'].append(loc)
            #####ROUTE
            
            route_lst = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['ROUTE'].dropna().drop_duplicates().to_list()
            route = '//'.join(map(str, route_lst))
            pm_master_dict['ROUTE'].append(route)
            
            #####LEADTIME
            pm_master_dict['LEADTIME']= leadtime
            #####PMCOUNTER
            pm_master_dict['PMCOUNTER']= ''
            #####WORKTYPE
            pm_master_dict['WORKTYPE']= worktype
            #####EGMNTACTTYPE
            pm_master_dict['EGMNTACTTYPE']= egmntacttype
            #####WOSTATUS
            pm_master_dict['WOSTATUS'] = wostatus
            ######EGCRAFT
            carft_lst = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['RESPONSE_CRAFT'].dropna().drop_duplicates().to_list()
            carft_lst = [str(x) for x in carft_lst] 
            carft = '//'.join(carft_lst)
            pm_master_dict['EGCRAFT'].append(carft)
            ######RESPONSED BY
            response_lst = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['RESPONSE'].dropna().drop_duplicates().to_list()
            response_lst = [str(x) for x in response_lst]
            response = '//'.join(response_lst)
            pm_master_dict['RESPONSED BY'].append(response)
            ######PTW
            ptw_lst = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['ประเภทของ_PERMIT_TO_WORK'].dropna().drop_duplicates().to_list()
            ptw_lst = [str(x) for x in ptw_lst]
            ptw = '//'.join(ptw_lst)
            pm_master_dict['PTW'].append(ptw)
            ######LOTO
            pm_master_dict['LOTO']= ''
            ######EGPROJECTID
            pm_master_dict['EGPROJECTID'] = egprojectid
            ######EGWBS
            pm_master_dict['EGWBS']= egwbs
            ######FREQUENCY
            pm_master_dict['FREQUENCY']= frequency
            ######FREQUNIT
            pm_master_dict['FREQUNIT']= frequnit
            ######NEXTDATE
            next_date = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['START_DATE'].min()
            pm_master_dict['NEXTDATE'].append(next_date.date())
            ######TARGSTARTTIME
            time_start = datetime.time(8,0)
            pm_master_dict['TARGSTARTTIME'] = time_start
            #######FINISH_DATE
            last_date = df_group_temp[df_group_temp['GROUP_LEVEL_1']==row['GROUP_LEVEL_1']]['FINISH_DATE'].max()
            pm_master_dict['FINISH_DATE'].append(last_date.date())
            #######FINISH TIME
            time_end = datetime.time(16,0)
            pm_master_dict['FINISH TIME'] = time_end
            #######PARENT
            pm_master_dict['PARENT']=''
            #######JPNUM
            JPNUM = 'JP'+'-'+row['GROUP_LEVEL_3']
            pm_master_dict['JPNUM'].append(JPNUM)
            #######main_system
            main_system = row['MAIN_SYSTEM']
            pm_master_dict['MAIN_SYSTEM'].append(main_system)
            #######sub_system
            sub_system = row['SUB_SYSTEM']
            pm_master_dict['SUB_SYSTEM'].append(sub_system)
            #######equipment
            equipment = row['EQUIPMENT']
            pm_master_dict['EQUIPMENT'].append(equipment)
            #######main_system_desc
            main_system_desc = row['MAIN_SYSTEM_DESC']
            pm_master_dict['MAIN_SYSTEM_DESC'].append(main_system_desc)
            #######sub_system_desc
            sub_system_desc = row['SUB_SYSTEM_DESC']
            pm_master_dict['SUB_SYSTEM_DESC'].append(sub_system_desc)
            #######kks_new_desc
            kks_new_desc = row['KKS_NEW_DESC']
            pm_master_dict['KKS_NEW_DESC'].append(kks_new_desc)
            #######unit
            unit = row['UNIT']
            pm_master_dict['UNIT'].append(unit)
            #######carft_desc
            TYPE = row['TYPE']
            pm_master_dict['TYPE'].append(TYPE)

        pm_master_df = pd.DataFrame(pm_master_dict)
        #? Clean '//'
        pm_master_df['PTW'] = pm_master_df['PTW'].apply(clean_ptw_column)
        return pm_master_df
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถสร้าง PM Plan ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>{str(e)}</li>"
            f"<li>เกิดข้อผิดพลาดในระหว่างการสร้าง PM Plan กรุณาตรวจสอบข้อมูลอีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def generate_pmnum(loop_num, type, unit, worktype):
    if worktype == 'APAO':
        return f'Group-{loop_num:02d}-{type}-{unit}-AD'
    else:
        return f'Group-{loop_num:02d}-{type}-{unit}'

def generate_description(*args):
    return ', '.join(map(str, args))

def group_pm_plan(request, pm_master_df, all_group_columns,
                    df_yyy_cont_more, df_yyy_cont_less, siteid,
                    group_columns, first_plant, location, worktype,
                    egmntacttype, wostatus, egprojectid, egwbs, frequency):
    try:
        #? PM PLAN MORE
        df_pm_plan3_more = pd.DataFrame()
        loop_num = 0
        for index,row in df_yyy_cont_more.iterrows():
            loop_num +=1 
            filter_condition = np.ones(len(pm_master_df), dtype=bool)
            
            for col in all_group_columns:
                filter_condition &= (pm_master_df[col] == row[col])

            group = pm_master_df[filter_condition]
            pm_parent_dict = {
                        'PMNUM':[],
                        'SITEID':[],
                        'DESCRIPTION':[],
                        'STATUS':[],
                        'LOCATION':[],
                        'ROUTE':[],
                        'LEADTIME':[],
                        'PMCOUNTER':[],
                        'WORKTYPE':[],
                        'EGMNTACTTYPE':[],
                        'WOSTATUS':[],
                        'EGCRAFT':[],
                        'RESPONSED BY':[],
                        'PTW':[],
                        'LOTO':[],
                        'EGPROJECTID':[],
                        'EGWBS':[],
                        'FREQUENCY':[],
                        'FREQUNIT':[],
                        'NEXTDATE':[],
                        'TARGSTARTTIME':[],
                        'FINISH_DATE':[],
                        'FINISH TIME':[],
                        'PARENT':[],
                        'JPNUM':[],
                        #'MAIN_SYSTEM':[],
                        #'SUB_SYSTEM':[],
                        #'EQUIPMENT':[],
                        #'MAIN_SYSTEM_DESC':[],
                        #'SUB_SYSTEM_DESC':[],
                        'UNIT':[],
                        'TYPE':[]
                        }
            pm_parent_dict['PMNUM'] = generate_pmnum(loop_num, row['TYPE'], row['UNIT'], worktype)
            pm_parent_dict['SITEID'] = siteid
            pm_parent_dict['DESCRIPTION'] = generate_description(*(row[col] for col in group_columns))
            pm_parent_dict['STATUS'] = status
            
            if 'MAIN_SYSTEM' in group_columns and 'MAIN_SYSTEM_DESC' in group_columns:
                pm_parent_dict['LOCATION'] = first_plant + row['MAIN_SYSTEM'] if pd.notna(row['MAIN_SYSTEM']) else location
            else:
                pm_parent_dict['LOCATION'] = first_plant + row['UNIT'] if pd.notna(row['UNIT']) else location
            
            pm_parent_dict['ROUTE'] = ''
            pm_parent_dict['LEADTIME'] = leadtime
            pm_parent_dict['PMCOUNTER'] =''
            pm_parent_dict['WORKTYPE'] = worktype
            pm_parent_dict['EGMNTACTTYPE'] = egmntacttype
            pm_parent_dict['WOSTATUS'] = wostatus
            pm_parent_dict['EGCRAFT'] = '_'.join(group['EGCRAFT'].dropna().drop_duplicates().to_list())
            pm_parent_dict['RESPONSED BY'] = '_'.join(group['RESPONSED BY'].dropna().drop_duplicates().to_list())
            pm_parent_dict['PTW'] = '_'.join(group['PTW'].dropna().drop_duplicates().to_list())
            pm_parent_dict['LOTO'] = ''
            pm_parent_dict['EGPROJECTID'] = egprojectid
            pm_parent_dict['EGWBS']  = egwbs
            pm_parent_dict['FREQUENCY'] = frequency
            pm_parent_dict['FREQUNIT'] = frequnit
            pm_parent_dict['NEXTDATE'] = group['NEXTDATE'].min() if not group['NEXTDATE'].isnull().all() else None
            pm_parent_dict['TARGSTARTTIME'] = group['TARGSTARTTIME'].min() 
            pm_parent_dict['FINISH_DATE'] = group['FINISH_DATE'].max() if not group['FINISH_DATE'].isnull().all() else None
            pm_parent_dict['FINISH TIME'] = group['FINISH TIME'].max()
            pm_parent_dict['PARENT'] = ''
            pm_parent_dict['JPNUM'] = ''
            # pm_parent_dict['MAIN_SYSTEM'] =row['MAIN_SYSTEM']
            # pm_parent_dict['SUB_SYSTEM'] =row['SUB_SYSTEM']
            # pm_parent_dict['EQUIPMENT'] =row['EQUIPMENT']
            # pm_parent_dict['MAIN_SYSTEM_DESC'] =row['MAIN_SYSTEM_DESC']
            # pm_parent_dict['SUB_SYSTEM_DESC'] =row['SUB_SYSTEM_DESC']
            pm_parent_dict['UNIT'] =row['UNIT']
            pm_parent_dict['TYPE'] =row['TYPE']
            parent_df = pd.DataFrame([pm_parent_dict])
            group.loc[:, 'PARENT'] = generate_pmnum(loop_num, row['TYPE'], row['UNIT'], worktype)
            parent = pd.concat([parent_df, group])
            parent.loc[:,'GROUP'] = loop_num
            df_pm_plan3_more = pd.concat([df_pm_plan3_more,parent])
        
        #? PM PLAN LESS
        df_pm_plan3_less = pd.DataFrame()
        loop_num = 0
        for index,row in df_yyy_cont_less.iterrows():
            loop_num +=1 
            filter_condition = np.ones(len(pm_master_df), dtype=bool)
            for col in all_group_columns:
                filter_condition &= (pm_master_df[col] == row[col])
                    
            group = pm_master_df[filter_condition]
            
            df_pm_plan3_less = pd.concat([df_pm_plan3_less, group])

        df_pm_plan3_more_ME = df_pm_plan3_more[df_pm_plan3_more['TYPE'] == 'ME']
        df_pm_plan3_more_EE = df_pm_plan3_more[df_pm_plan3_more['TYPE'] == 'EE']
        df_pm_plan3_more_CV = df_pm_plan3_more[df_pm_plan3_more['TYPE'] == 'CV']
        df_pm_plan3_more_IC = df_pm_plan3_more[df_pm_plan3_more['TYPE'] == 'IC']

        df_pm_plan3_less_ME = df_pm_plan3_less[df_pm_plan3_less['TYPE'] == 'ME']
        df_pm_plan3_less_EE = df_pm_plan3_less[df_pm_plan3_less['TYPE'] == 'EE']
        df_pm_plan3_less_CV = df_pm_plan3_less[df_pm_plan3_less['TYPE'] == 'CV']
        df_pm_plan3_less_IC = df_pm_plan3_less[df_pm_plan3_less['TYPE'] == 'IC']

        lst_group  = [df_pm_plan3_more_ME,df_pm_plan3_more_EE,df_pm_plan3_more_CV,df_pm_plan3_more_IC]
        lst_no_group = [df_pm_plan3_less_ME,df_pm_plan3_less_EE, df_pm_plan3_less_CV,df_pm_plan3_less_IC]

        df_pm_plan3_test = pd.DataFrame()
        for i in range(len(lst_group)):
            df_pm_plan3_test = pd.concat([df_pm_plan3_test,lst_group[i]])
            df_pm_plan3_test = pd.concat([df_pm_plan3_test,lst_no_group[i]])

        df_pm_plan3_test['MOD'] = df_pm_plan3_test['GROUP']%2
        df_pm_plan3 = df_pm_plan3_test.copy()
        return df_pm_plan3
    except KeyError as e:
        logger.error(f"Key error: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบข้อมูลที่จำเป็นสำหรับการดำเนินการ<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>{str(e)}</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบว่าชื่อคอลัมน์ถูกต้องและครบถ้วน  หากยังพบปัญหา โปรดติดต่อทีมสนับสนุนเพื่อขอความช่วยเหลือ</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถจัดกลุ่ม PM Plan ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>{str(e)}</li>"
            f"<li>เกิดข้อผิดพลาดในระหว่างการจัดกลุ่มข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def replace_columns(col, replace_dict):
    try:
        for key, value in replace_dict.items():
            col = re.sub(rf'\b{key}\b', value, col)
    except Exception as e:
        logger.error(f"An error occurred: {e}")
    return col

def create_workorder(request, df, pmnum_col, orgid):
    try:
        df_workorder = df[['SITEID', pmnum_col, 'WOSTATUS', 'NEXTDATE', 'FINISH_DATE']].copy()
        df_workorder['PARENTCHGSSTATUS'] = 0
        df_workorder['ORGID'] = orgid
        df_workorder['NEXTDATE'] = df_workorder['NEXTDATE'] + ' ' + '01:00:00'
        df_workorder['FINISH_DATE'] = df_workorder['FINISH_DATE'] + ' ' + '09:00:00'
        df_workorder['SCHEDSTART'] = df_workorder['NEXTDATE']
        df_workorder['SCHEDFINISH'] = df_workorder['FINISH_DATE']
        df_workorder.rename(columns={'WOSTATUS': 'STATUS', 'NEXTDATE': 'TARGSTARTDATE', 'FINISH_DATE': 'TARGCOMPDATE'}, inplace=True)
        columns_order = ['SITEID', 'ORGID', 'STATUS', pmnum_col, 'PARENTCHGSSTATUS', 'TARGSTARTDATE', 'TARGCOMPDATE', 'SCHEDSTART', 'SCHEDFINISH']
        df_workorder = df_workorder[columns_order]
        
        return df_workorder

    except Exception as e:
        logger.error(f"Failed to create workorder DataFrame: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถสร้างข้อมูล Workorder ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดในระหว่างการสร้างข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def copy_worksheet(request, file_path, sheet_name, index):
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet name '{sheet_name}' does not exist in the workbook.")
        source_sheet = wb[sheet_name]
        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = f"{sheet_name}-{index}"
        wb.save(file_path)
    
    except FileNotFoundError:
        logger.error(f"Excel file '{file_path}' not found.", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบไฟล์ Excel '{file_path}' ที่ระบุ<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>ระบบไม่สามารถหาไฟล์ที่ระบุได้ กรุณาตรวจสอบเส้นทางไฟล์และลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

    except ValueError as ve:
        logger.error(f"Invalid sheet name '{sheet_name}': {ve}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่พบชีท '{sheet_name}' ในไฟล์ Excel ที่ระบุ<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>ชื่อชีทที่ระบุไม่ถูกต้องหรือไม่มีอยู่ในไฟล์</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: กรุณาตรวจสอบชื่อชีทและลองอีกครั้ง หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือ</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

    except Exception as e:
        logger.error(f"Failed to copy worksheet '{sheet_name}' in file '{file_path}': {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถคัดลอกชีท '{sheet_name}' ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดระหว่างการคัดลอกชีท กรุณาตรวจสอบข้อมูลและลองใหม่อีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def write_dataframes_to_excel(request, writer, df_labor, lst_labor, df_jop, jop_plan_column, df_pm, pm_plan_column, df_workorder, sheet_name_labor, sheet_name_task, sheet_name_pm, sheet_name_wo, start_row, start_offset):
    try:
        df_labor[lst_labor].to_excel(writer, sheet_name=sheet_name_labor, startrow=start_row, startcol=0, index=False, header=False)
        df_jop[jop_plan_column].to_excel(writer, sheet_name=sheet_name_task, startrow=start_row, startcol=0, index=False, header=False)
        df_pm[pm_plan_column].to_excel(writer, sheet_name=sheet_name_pm, startrow=start_row, startcol=0, index=False, header=False)
        df_workorder.to_excel(writer, sheet_name=sheet_name_wo, startrow=(start_row + start_offset), startcol=1, index=False, header=False)
    except Exception as e:
        logger.error(f"Failed to write dataframes to Excel: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถเขียนข้อมูลลงในไฟล์ Excel ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดในระหว่างการเขียนข้อมูล กรุณาตรวจสอบข้อมูลอีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')

def decorate_sheet_labor(sheet, thin_border, fill_color, yellow_fill):
    # ตกแต่งชีต JPPLAN-LABOR
    headers = ['GROUP', 'MOD', 'COMMENT']
    for i, header in enumerate(headers):
        cell = sheet.cell(row=2, column=11 + i, value=header)
        cell.border = thin_border

    for row in range(3, sheet.max_row + 1):
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border

            if sheet.cell(row=row, column=12).value == 0:
                sheet.cell(row=row, column=col).fill = fill_color

            if sheet.cell(row=row, column=13).value:
                sheet.cell(row=row, column=col).fill = yellow_fill

    sheet.insert_cols(idx=11)
    sheet.auto_filter.ref = "A2:N2"

def decorate_sheet_task(sheet, thin_border, fill_color, yellow_fill, start_row, df_jop_plan_master):
    headers = ['START_DATE', 'FINISH_DATE', 'GROUP', 'MOD', 'COMMENT']
    for row in range(start_row + 1, start_row + 1 + len(df_jop_plan_master)):
        sheet[f'L{row}'].number_format = 'YYYY-MM-DD'
        sheet[f'M{row}'].number_format = 'YYYY-MM-DD'
    
    for i, header in enumerate(headers):
        cell = sheet.cell(row=2, column=12 + i, value=header)
        cell.border = thin_border

    for row in range(3, sheet.max_row + 1):
        for col in range(1, 12):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border

            if sheet.cell(row=row, column=15).value == 0:
                sheet.cell(row=row, column=col).fill = fill_color

            if sheet.cell(row=row, column=16).value:
                sheet.cell(row=row, column=col).fill = yellow_fill

    sheet.insert_cols(idx=12)
    sheet.column_dimensions['M'].width = 20.0
    sheet.column_dimensions['N'].width = 20.0
    sheet.auto_filter.ref = "A2:Q2"

def decorate_sheet_pm(sheet, thin_border, fill_color, yellow_fill, start_row, df_pm_plan3_master):
    headers = ['MOD', 'MAIN_SYSTEM', 'MAIN_SYSTEM_DESC', 'UNIT',
            'TYPE','SUB_SYSTEM', 'EQUIPMENT', 'SUB_SYSTEM_DESC',
            'KKS_NEW_DESC', 'GROUP', 'FINISH_DATE', 'FINISH TIME', 'COMMENT']
    
    for row in range(start_row + 1, start_row + 1 + len(df_pm_plan3_master)):
        sheet[f'O{row}'].number_format = 'YYYY-MM-DD'
        sheet[f'U{row}'].number_format = 'hh:mm:ss'
        sheet[f'AI{row}'].number_format = 'YYYY-MM-DD'
    
    for i, header in enumerate(headers):
        cell = sheet.cell(row = 2, column = 25 + i, value = header)
        cell.border = thin_border

    for row in range(3, sheet.max_row + 1):
        for col in range(1,25):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            
            if sheet.cell(row=row, column=25).value == 0:
                sheet.cell(row=row, column=col).fill = fill_color

            if not sheet.cell(row=row, column=13).value:
                sheet.cell(row=row, column=col).font = Font(bold=True)
                
            if sheet.cell(row=row, column=37).value:
                sheet.cell(row=row, column=col).fill = yellow_fill

    sheet.insert_cols(idx=25)
    sheet.column_dimensions['AA'].width = 15.0
    # sheet.column_dimensions['AB'].width = 51.0
    sheet.column_dimensions['AE'].width = 15.0
    # sheet.column_dimensions['AG'].width = 51.0
    # sheet.column_dimensions['AH'].width = 90.0
    sheet.column_dimensions['AJ'].width = 15.0
    sheet.column_dimensions['AK'].width = 15.0
    sheet.auto_filter.ref = 'A2:AL2'

def decorate_sheet(sheet, sheet_name, thin_border, fill_color, yellow_fill, start_row, df_jop_plan_master, df_pm_plan3_master):
    if "JPPLAN-LABOR" in sheet_name:
        decorate_sheet_labor(sheet, thin_border, fill_color, yellow_fill)
    elif "JPPLAN-TASK" in sheet_name:
        decorate_sheet_task(sheet, thin_border, fill_color, yellow_fill, start_row, df_jop_plan_master)  # ส่ง df_jop_plan_master มาด้วย
    elif "PMPlan" in sheet_name:
        decorate_sheet_pm(sheet, thin_border, fill_color, yellow_fill, start_row, df_pm_plan3_master)
    # elif "WO" in sheet_name:
    #     decorate_sheet_wo(sheet)

def copy_sheets_to_macro_file(request, file_template_xlsx, file_path, file_template_xlsm, basic_sheet_names, location):
    try:
        with xw.App(visible=False) as app:
            time.sleep(2)

            book_xlsx = app.books.open(file_template_xlsx)
            book_xlsm = app.books.open(file_path)  # เปิดไฟล์ต้นฉบับ

            # ลบชีตเดิมใน .xlsm ที่ต้องการแทนที่
            for sheet_name in basic_sheet_names:
                if sheet_name in [s.name for s in book_xlsm.sheets]:
                    book_xlsm.sheets[sheet_name].delete()
            
            # คัดลอกชีตทั้งหมดจาก .xlsx ไปยัง .xlsm
            for sheet_name in basic_sheet_names:
                sheet_xlsx = book_xlsx.sheets[sheet_name]
                sheet_xlsx.api.Copy(After=book_xlsm.sheets[-1].api) # คัดลอกชีตจาก .xlsx ไปยัง .xlsm วางไว้หลังชีตสุดท้ายใน .xlsm

                new_sheet = book_xlsm.sheets[-1]  # ชีตที่ถูกคัดลอกจะเป็นชีตสุดท้าย
                if sheet_name in [s.name for s in book_xlsm.sheets]:
                    new_sheet_name = f"{sheet_name}-{location}"
                else:
                    new_sheet_name = sheet_name

                new_sheet.name = new_sheet_name
                
            for sheet_name in basic_sheet_names:
                sheet_name_with_suffix = f"{sheet_name}-1"
                sheet_xlsx = book_xlsx.sheets[sheet_name_with_suffix]
                sheet_xlsx.api.Copy(After=book_xlsm.sheets[-1].api)

                new_sheet = book_xlsm.sheets[-1]
                if sheet_name_with_suffix in [s.name for s in book_xlsm.sheets]:
                    new_sheet_name = f"{sheet_name_with_suffix}-{location}"
                else:
                    new_sheet_name = sheet_name_with_suffix

                new_sheet.name = new_sheet_name

            book_xlsm.save(file_template_xlsm)
            print(f"File saved successfully as {file_template_xlsm}")
            
    except Exception as e:
        logger.error(f"Failed to copy sheets to Excel: {str(e)}", exc_info=True)
        error_message = (
            f"<div class='error-container'>"
            f"<strong class='error-title'>พบปัญหา:</strong> ไม่สามารถคัดลอกชีตไปยังไฟล์ Excel ได้<br>"
            f"<ul class='error-details'>"
            f"<p class='error-description'>สาเหตุของปัญหา:</p>"
            f"<li>เกิดข้อผิดพลาดในระหว่างคัดลอกชีต กรุณาตรวจสอบข้อมูลอีกครั้ง</li>"
            f"</ul>"
            f"<p class='error-note'>คำแนะนำ: หากยังพบปัญหา โปรดติดต่อฝ่ายสนับสนุนเพื่อขอความช่วยเหลือเพิ่มเติม</p>"
            f"</div>"
        )
        messages.error(request, error_message)
        return redirect('index')